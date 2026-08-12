"""
==============================================
MACRO ECONOMIC ANALYSIS - EXCEL REPORT GENERATOR
==============================================
Tạo file Excel phân tích tương quan giữa:
- Lượt khách nhà hàng (từ DB thực)
- Giá vàng SJC
- Giá xăng RON95
- Lãi suất huy động Vietcombank kỳ hạn 12 tháng (tại quầy)
- Lãi suất cho vay bình quân

Deposit Rate Source: Vietcombank (vietcombank.com.vn)
  Ref: CafeF, Lao Dong, DNSE, webgia.com

Output: Macro_Economic_Analysis.xlsx (đa sheet, có chart, conditional formatting)
"""

import sys
import os
import datetime
import traceback

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

from forecast_system.config.settings import PROJECT_ROOT, CURRENT_DATE
from forecast_system.utils.db_utils import create_db_engine
from forecast_system.utils.logger import get_logger

logger = get_logger('macro_analysis')

OUTPUT_FILE = str(PROJECT_ROOT / "Macro_Economic_Analysis.xlsx")

# ==========================================
# STYLES
# ==========================================
DARK_BLUE = "1B2A4A"
MEDIUM_BLUE = "2B4C7E"
ACCENT_BLUE = "4472C4"
LIGHT_BLUE = "D6E4F0"
HEADER_GOLD = "E8B931"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
RED_ALERT = "FF4444"
GREEN_GOOD = "27AE60"
ORANGE_WARN = "F39C12"

THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)

HEADER_FONT = Font(name='Calibri', bold=True, color=WHITE, size=11)
HEADER_FILL = PatternFill('solid', fgColor=DARK_BLUE)
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

SUBHEADER_FONT = Font(name='Calibri', bold=True, color=DARK_BLUE, size=10)
SUBHEADER_FILL = PatternFill('solid', fgColor=LIGHT_BLUE)

TITLE_FONT = Font(name='Calibri', bold=True, color=DARK_BLUE, size=16)
SUBTITLE_FONT = Font(name='Calibri', bold=False, color=MEDIUM_BLUE, size=12)

DATA_FONT = Font(name='Calibri', size=10)
DATA_ALIGN = Alignment(horizontal='center', vertical='center')
DATA_ALIGN_LEFT = Alignment(horizontal='left', vertical='center')

GOLD_FILL = PatternFill('solid', fgColor='FFF2CC')
GREEN_FILL = PatternFill('solid', fgColor='E2EFDA')
RED_FILL = PatternFill('solid', fgColor='FCE4EC')
ORANGE_FILL = PatternFill('solid', fgColor='FFF3E0')


# ==========================================
# MACRO ECONOMIC DATA (Curated Monthly)
# ==========================================
def generate_macro_data():
    """
    Dữ liệu kinh tế vĩ mô Việt Nam theo tháng (2022-01 → 2026-03).
    Sources: SBV, SJC, Petrolimex, GSO.
    Deposit Rate: Vietcombank kỳ hạn 12 tháng (tại quầy).
      Ref: vietcombank.com.vn, CafeF, Lao Dong, DNSE, webgia.com
    """
    records = []
    
    # (year, month, gold_sjc_mil, gas_ron95_kd, ls_vcb_12m_pct, ls_cho_vay_bq_pct)
    # ls_vcb_12m_pct = Lãi suất Vietcombank kỳ hạn 12 tháng (tại quầy)
    data_raw = [
        # 2022 — VCB ổn định 5.3-5.5% trước Q3, tăng mạnh Q4 theo SBV
        (2022, 1,  62.0, 23.2, 5.50, 9.2),
        (2022, 2,  63.5, 24.5, 5.50, 9.2),
        (2022, 3,  68.0, 28.0, 5.50, 9.1),
        (2022, 4,  69.5, 29.5, 5.40, 9.0),
        (2022, 5,  69.0, 30.0, 5.30, 8.9),
        (2022, 6,  67.5, 31.5, 5.30, 8.8),
        (2022, 7,  66.0, 30.0, 5.50, 8.7),
        (2022, 8,  66.5, 28.5, 5.50, 8.7),
        (2022, 9,  66.0, 27.0, 6.40, 8.8),  # VCB bắt đầu tăng mạnh
        (2022, 10, 66.5, 27.5, 6.80, 8.9),
        (2022, 11, 67.0, 26.5, 7.10, 9.3),
        (2022, 12, 67.5, 25.0, 7.40, 9.5),  # VCB đỉnh cuối 2022
        # 2023 — VCB giữ đỉnh Q1, giảm mạnh từ Q2 theo SBV nới lỏng
        (2023, 1,  67.0, 24.0, 7.40, 9.6),  # VCB đỉnh chu kỳ
        (2023, 2,  67.5, 23.5, 7.40, 9.5),
        (2023, 3,  68.0, 23.0, 7.40, 9.4),
        (2023, 4,  68.5, 22.5, 6.80, 9.2),  # VCB bắt đầu giảm
        (2023, 5,  67.5, 21.5, 6.50, 9.0),
        (2023, 6,  67.0, 21.0, 6.30, 8.8),
        (2023, 7,  67.5, 22.0, 5.80, 8.6),
        (2023, 8,  68.0, 23.0, 5.50, 8.5),
        (2023, 9,  69.0, 24.0, 5.30, 8.5),
        (2023, 10, 70.5, 23.5, 5.10, 8.5),
        (2023, 11, 71.0, 22.5, 4.90, 8.5),
        (2023, 12, 73.0, 22.0, 4.80, 8.5),  # VCB đáy Q4/2023
        # 2024 — VCB đáy lịch sử 4.6%, ổn định cả năm
        (2024, 1,  74.2, 21.9, 4.70, 8.5),
        (2024, 2,  76.5, 22.3, 4.70, 8.4),
        (2024, 3,  80.0, 22.8, 4.60, 8.3),  # VCB chạm đáy 4.6%
        (2024, 4,  84.5, 23.5, 4.60, 8.1),
        (2024, 5,  90.0, 23.0, 4.60, 7.9),
        (2024, 6,  85.2, 22.5, 4.60, 7.7),
        (2024, 7,  82.0, 22.0, 4.60, 7.5),
        (2024, 8,  81.0, 21.5, 4.60, 7.3),
        (2024, 9,  84.0, 21.8, 4.60, 7.2),
        (2024, 10, 87.5, 22.0, 4.70, 7.1),
        (2024, 11, 92.0, 22.5, 4.70, 7.0),
        (2024, 12, 95.0, 23.0, 4.70, 6.9),  # VCB ổn định ~4.7%
        # 2025 — VCB duy trì thấp 4.6-4.8%
        (2025, 1, 100.0, 23.5, 4.70, 6.9),
        (2025, 2, 105.0, 24.0, 4.70, 6.8),
        (2025, 3, 108.0, 24.5, 4.70, 6.8),
        (2025, 4, 112.0, 25.0, 4.60, 6.8),
        (2025, 5, 118.0, 25.5, 4.60, 6.7),
        (2025, 6, 125.0, 26.0, 4.60, 6.7),
        (2025, 7, 130.0, 26.5, 4.60, 6.7),
        (2025, 8, 135.0, 27.0, 4.70, 6.7),
        (2025, 9, 140.0, 27.5, 4.70, 6.6),
        (2025, 10, 148.0, 28.0, 4.70, 6.6),
        (2025, 11, 155.0, 28.5, 4.70, 6.6),
        (2025, 12, 160.0, 29.0, 4.80, 6.6),  # VCB nhích nhẹ cuối 2025
        # 2026 — VCB tăng mạnh Q1/2026
        (2026, 1, 165.0, 29.5, 5.20, 6.6),  # VCB bắt đầu tăng
        (2026, 2, 170.0, 30.0, 5.50, 6.6),
        (2026, 3, 175.0, 30.0, 5.90, 6.6),  # VCB tăng mạnh Q1/2026
    ]
    
    for y, m, gold, gas, ls_hd, ls_cv in data_raw:
        dt = datetime.date(y, m, 1)
        records.append({
            'Year': y,
            'Month': m,
            'Date': dt,
            'Date_Label': dt.strftime('%m/%Y'),
            'Gia_Vang_SJC_Trieu': gold,
            'Gia_Xang_RON95_KD': gas,
            'LS_Huy_Dong_12T_Pct': ls_hd,
            'LS_Cho_Vay_BQ_Pct': ls_cv,
        })
    
    return pd.DataFrame(records)



# ╔══════════════════════════════════════════════════════════════════════╗
# ║  ⚠️ CRITICAL DEDUPLICATION LOGIC — DO NOT MODIFY OR REMOVE ⚠️      ║
# ║                                                                     ║
# ║  order_guid (payment_hub) có dạng: '{XXXXXXXX-XXXX-XXXX-XXXX}'    ║
# ║  transaction_id (rk_dc)   có dạng: 'XXXXXXXX-XXXX-XXXX-XXXX'     ║
# ║                                                                     ║
# ║  → PHẢI dùng REPLACE để loại bỏ ký tự '{' và '}' trước khi so    ║
# ║    sánh, nếu không sẽ đếm TRÙNG gấp đôi (cùng 1 giao dịch       ║
# ║    xuất hiện ở cả 2 bảng nhưng txn_id không khớp do dấu {}).     ║
# ║                                                                     ║
# ║  → Dùng ROW_NUMBER(PARTITION BY clean_txn_id) để giữ duy nhất     ║
# ║    1 bản ghi cho mỗi giao dịch, ưu tiên payment_hub.             ║
# ║                                                                     ║
# ║  VÍ DỤ THỰC TẾ: T06/2024 payment_hub = 1,768,309 khách.         ║
# ║  Nếu KHÔNG strip {} → kết quả sai = 3,597,947 (gấp đôi).        ║
# ╚══════════════════════════════════════════════════════════════════════╝
QUERY_MONTHLY_DEDUPLICATED = """
    WITH all_txns AS (
        SELECT 
            REPLACE(REPLACE(order_guid, '{', ''), '}', '') AS txn_id,
            shift_date,
            COALESCE(guest_count, 0) AS guest_count,
            COALESCE(total_pay_sum, 0) AS total_revenue,
            restaurant_code,
            1 AS source_priority
        FROM v_fact_db_payment_hub_transactions
        WHERE shift_date >= :start_date AND shift_date < :end_date

        UNION ALL

        SELECT
            REPLACE(REPLACE(transaction_id, '{', ''), '}', '') AS txn_id,
            shiftdate AS shift_date,
            COALESCE(guest_count, 0) AS guest_count,
            COALESCE(paysum, 0) AS total_revenue,
            restaurant_code,
            2 AS source_priority
        FROM v_fact_db_rk_dc_transactions
        WHERE shiftdate >= :start_date AND shiftdate < :end_date
    ),
    deduped AS (
        SELECT txn_id, shift_date, guest_count, total_revenue, restaurant_code,
            ROW_NUMBER() OVER (PARTITION BY txn_id ORDER BY source_priority) AS rn
        FROM all_txns
    )
    SELECT
        YEAR(shift_date) AS Year,
        MONTH(shift_date) AS Month,
        SUM(guest_count) AS Total_Guests,
        SUM(total_revenue) AS Total_Revenue,
        COUNT(DISTINCT txn_id) AS Total_Transactions,
        COUNT(DISTINCT restaurant_code) AS Active_Restaurants,
        COUNT(DISTINCT shift_date) AS Active_Days
    FROM deduped
    WHERE rn = 1
    GROUP BY YEAR(shift_date), MONTH(shift_date)
    ORDER BY Year, Month
"""


# ==========================================
# LOAD 2022 DATA FROM SALES_BY_STORE.XLSX (Priority source for 2022)
# ==========================================
SALES_BY_STORE_FILE = str(PROJECT_ROOT / "Sales_by_Store.xlsx")

def load_guest_data_from_sales_by_store():
    """
    Load dữ liệu 2022 từ Sales_by_Store.xlsx.
    - Cột 'Sales Act' → Total_Revenue (doanh thu thực tế)
    - Cột 'TC' → Total_Guests (lượt khách / transaction count)
    """
    import os
    if not os.path.exists(SALES_BY_STORE_FILE):
        logger.warning(f"Sales_by_Store.xlsx not found: {SALES_BY_STORE_FILE}")
        return pd.DataFrame()
    
    logger.info(f"📂 Loading 2022 data from: {SALES_BY_STORE_FILE}")
    print(f"   → Reading Sales_by_Store.xlsx...", end="", flush=True)
    
    try:
        import warnings
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            df = pd.read_excel(SALES_BY_STORE_FILE, sheet_name='Export')
    except Exception as e:
        logger.error(f"Failed to read Sales_by_Store.xlsx: {e}")
        return pd.DataFrame()
    
    # Parse dates
    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
    df = df[df['Date'].notna()].copy()
    
    # Parse numeric columns
    df['Sales Act'] = pd.to_numeric(df['Sales Act'], errors='coerce').fillna(0)
    df['TC'] = pd.to_numeric(df['TC'], errors='coerce').fillna(0)
    
    # Filter only positive data
    df = df[(df['Sales Act'] > 0) | (df['TC'] > 0)].copy()
    
    if df.empty:
        logger.warning("No valid data in Sales_by_Store.xlsx")
        return pd.DataFrame()
    
    df['Year'] = df['Date'].dt.year
    df['Month'] = df['Date'].dt.month
    
    # Aggregate by month
    monthly = df.groupby(['Year', 'Month']).agg(
        Total_Revenue=('Sales Act', 'sum'),
        Total_Guests=('TC', 'sum'),
        Total_Transactions=('TC', 'count'),  # number of rows as proxy
        Active_Restaurants=('Store', 'nunique'),
        Active_Days=('Date', 'nunique'),
    ).reset_index()
    
    # Ensure numeric
    for col in ['Total_Guests', 'Total_Revenue', 'Total_Transactions', 'Active_Restaurants', 'Active_Days']:
        monthly[col] = pd.to_numeric(monthly[col], errors='coerce').fillna(0)
    
    # Filter to only full months (>= 25 active days) to exclude partial months like T03/2022 (9 days)
    full_months = monthly[monthly['Active_Days'] >= 25].copy()
    partial = monthly[monthly['Active_Days'] < 25]
    if not partial.empty:
        labels = list(partial.apply(
            lambda r: f"{int(r['Year'])}/{int(r['Month']):02d} ({int(r['Active_Days'])}d)", axis=1
        ))
        logger.info(f"   ⚠️  Skipped {len(partial)} partial months from Sales_by_Store: {labels}")
        print(f" ({len(partial)} partial months skipped)", end="")
    
    # Derived metrics
    full_months['Avg_Daily_Guests'] = (
        full_months['Total_Guests'] / full_months['Active_Days'].replace(0, np.nan)
    ).round(0)
    full_months['Avg_Guests_Per_Restaurant'] = (
        full_months['Total_Guests'] / full_months['Active_Restaurants'].replace(0, np.nan)
        / full_months['Active_Days'].replace(0, np.nan)
    ).round(1)
    full_months['Avg_Revenue_Per_Guest'] = (
        full_months['Total_Revenue'] / full_months['Total_Guests'].replace(0, np.nan)
    ).round(0)
    full_months['Avg_Daily_Revenue'] = (
        full_months['Total_Revenue'] / full_months['Active_Days'].replace(0, np.nan)
    ).round(0)
    
    total_rev = full_months['Total_Revenue'].sum()
    total_guests = full_months['Total_Guests'].sum()
    print(f" ✅ {len(full_months)} months")
    print(f"   → Date range: {df['Date'].min().date()} → {df['Date'].max().date()}")
    print(f"   → Total: {total_guests:,.0f} guests | {total_rev:,.0f} VNĐ revenue")
    
    logger.info(f"✅ Sales_by_Store: {len(full_months)} months | {total_guests:,.0f} guests | {total_rev:,.0f} VNĐ")
    return full_months


def load_guest_data_from_db():
    """
    Load dữ liệu khách + doanh thu từ DB (payment_hub + rk_dc).
    ⚠️ DEDUPLICATION: Cùng 1 transaction_id có thể xuất hiện ở CẢ HAI bảng.
    Query dùng CTE + ROW_NUMBER(PARTITION BY txn_id) để giữ duy nhất 1 bản ghi
    cho mỗi transaction, ưu tiên payment_hub.
    Aggregates at SQL level → returns ~30 rows, not millions.
    """
    logger.info("🔌 Connecting to database...")
    engine = create_db_engine()
    if engine is None:
        logger.error("Cannot connect to DB")
        return pd.DataFrame()
    
    end_dt = CURRENT_DATE + datetime.timedelta(days=1)
    params = {
        'start_date': '2022-01-01',  # Start from 2022 to match macro data range
        'end_date': end_dt.strftime('%Y-%m-%d'),
    }
    
    logger.info(f"📥 Querying monthly aggregates from DB (up to {CURRENT_DATE})...")
    print("   → Using DEDUPLICATED SQL query (UNION ALL + ROW_NUMBER)...")
    print("   → Dedup: payment_hub ∪ rk_dc, loại trùng theo txn_id...", end="", flush=True)
    
    try:
        from sqlalchemy import text as sa_text
        
        with engine.connect() as conn:
            monthly = pd.read_sql(sa_text(QUERY_MONTHLY_DEDUPLICATED), conn, params=params)
            print(f" ✅ {len(monthly)} months")
    
    except Exception as e:
        logger.error(f"DB query failed: {e}")
        traceback.print_exc()
        engine.dispose()
        return pd.DataFrame()
    
    if monthly.empty:
        logger.warning("No data returned from DB (deduplicated query)")
        engine.dispose()
        return pd.DataFrame()
    
    # Ensure numeric
    for col in ['Total_Guests', 'Total_Revenue', 'Total_Transactions', 'Active_Restaurants', 'Active_Days']:
        monthly[col] = pd.to_numeric(monthly[col], errors='coerce').fillna(0)
    
    # Derived metrics
    monthly['Avg_Daily_Guests'] = (monthly['Total_Guests'] / monthly['Active_Days'].replace(0, np.nan)).round(0)
    monthly['Avg_Guests_Per_Restaurant'] = (
        monthly['Total_Guests'] / monthly['Active_Restaurants'].replace(0, np.nan) / monthly['Active_Days'].replace(0, np.nan)
    ).round(1)
    monthly['Avg_Revenue_Per_Guest'] = (monthly['Total_Revenue'] / monthly['Total_Guests'].replace(0, np.nan)).round(0)
    monthly['Avg_Daily_Revenue'] = (monthly['Total_Revenue'] / monthly['Active_Days'].replace(0, np.nan)).round(0)
    
    # Filter to only full months (>= 25 active days), except current month
    current_ym = (CURRENT_DATE.year, CURRENT_DATE.month)
    full_months = monthly[
        (monthly['Active_Days'] >= 25) | 
        ((monthly['Year'] == current_ym[0]) & (monthly['Month'] == current_ym[1]))
    ].copy()
    
    skipped = monthly[~monthly.index.isin(full_months.index)]
    if not skipped.empty:
        labels = list(skipped.apply(lambda r: f"{int(r['Year'])}/{int(r['Month']):02d} ({int(r['Active_Days'])}d)", axis=1))
        logger.info(f"   ⚠️  Skipped {len(skipped)} partial months: {labels}")
    
    total_rev = full_months['Total_Revenue'].sum()
    total_guests = full_months['Total_Guests'].sum()
    date_range = f"{int(full_months['Year'].min())}/{int(full_months['Month'].min()):02d} → {int(full_months['Year'].max())}/{int(full_months['Month'].max()):02d}"
    
    print(f"   → {len(full_months)} full months loaded ({date_range})")
    print(f"   → Total: {total_guests:,.0f} guests | {total_rev:,.0f} VNĐ revenue")
    logger.info(f"✅ DB loaded: {len(full_months)} months | {total_guests:,.0f} guests | {total_rev:,.0f} VNĐ")
    
    engine.dispose()
    return full_months


def calculate_correlations(df_merged):
    """Tính Pearson correlation giữa các yếu tố vĩ mô và lượt khách + doanh thu."""
    macro_cols = [
        'Gia_Vang_SJC_Trieu', 'Gia_Xang_RON95_KD',
        'LS_Huy_Dong_12T_Pct', 'LS_Cho_Vay_BQ_Pct'
    ]
    target_cols = [
        ('Total_Guests', 'Lượt Khách'),
        ('Total_Revenue', 'Doanh Thu'),
    ]
    
    results = []
    for target_col, target_label in target_cols:
        if target_col not in df_merged.columns:
            continue
        # Skip revenue correlations if no revenue data (CSV fallback)
        if target_col == 'Total_Revenue' and df_merged[target_col].sum() == 0:
            logger.info(f"   ⚠️  Skipping '{target_label}' correlations — no revenue data in this source")
            continue
        for col in macro_cols:
            valid = df_merged[[col, target_col]].dropna()
            if len(valid) >= 3:
                corr = valid[col].corr(valid[target_col])
                
                if pd.isna(corr):
                    results.append({
                        'Target': target_label,
                        'Target_Col': target_col,
                        'Yeu_To': col,
                        'He_So_r': 0.0,
                        'Abs_r': 0.0,
                        'Huong': 'N/A',
                        'Muc_Do': 'Không tính được',
                        'Y_Nghia': f'Giá trị không đổi trong kỳ → không thể tính tương quan',
                    })
                    continue
                
                abs_r = abs(corr)
                if abs_r >= 0.7:
                    strength = "Mạnh"
                elif abs_r >= 0.4:
                    strength = "Trung bình"
                elif abs_r >= 0.2:
                    strength = "Yếu"
                else:
                    strength = "Không đáng kể"
                
                direction = "Thuận (+)" if corr > 0 else "Nghịch (-)"
                
                results.append({
                    'Target': target_label,
                    'Target_Col': target_col,
                    'Yeu_To': col,
                    'He_So_r': round(corr, 4),
                    'Abs_r': round(abs_r, 4),
                    'Huong': direction,
                    'Muc_Do': strength,
                    'Y_Nghia': _interpret_correlation(col, corr, target_label),
                })
    
    if not results:
        return pd.DataFrame()
    return pd.DataFrame(results).sort_values(['Target', 'Abs_r'], ascending=[False, False])


def _interpret_correlation(col_name, r, target='Lượt Khách'):
    """Diễn giải ý nghĩa tương quan."""
    t = target.lower()
    if 'Vang' in col_name:
        if r > 0.4:
            return f"Vàng tăng, {target.lower()} vẫn tăng — F&B không bị ảnh hưởng trực tiếp bởi giá vàng"
        elif r < -0.4:
            return f"Vàng tăng → {target.lower()} giảm: người dân dồn tiền vào vàng, cắt chi tiêu ăn uống"
        return f"Giá vàng ít tác động đến {target.lower()} nhà hàng"
    elif 'Xang' in col_name:
        if r < -0.4:
            return f"Xăng đắt → chi phí tăng → {target.lower()} giảm"
        elif r > 0.4:
            return f"Xăng tăng cùng xu hướng kinh tế tăng trưởng → {target.lower()} vẫn tăng"
        return f"Giá xăng có tác động vừa phải đến {target.lower()}"
    elif 'Huy_Dong' in col_name:
        if r > 0.4:
            return f"LS VCB 12T cao → tiền gửi tiết kiệm nhiều hơn, nhưng {target.lower()} vẫn ổn"
        elif r < -0.4:
            return f"LS VCB 12T cao → dân gửi tiết kiệm thay vì tiêu dùng, {target.lower()} giảm"
        return f"LS VCB 12T ít ảnh hưởng trực tiếp đến {target.lower()}"
    elif 'Cho_Vay' in col_name:
        if r < -0.4:
            return f"LS cho vay cao → DN + cá nhân thắt chặt chi tiêu → {target.lower()} giảm"
        elif r > 0.4:
            return f"LS cho vay tương quan thuận với {target.lower()} — xu hướng đồng thời tăng"
        return f"LS cho vay có tác động vừa phải lên {target.lower()}"
    return ""


# ==========================================
# EXCEL BUILDER
# ==========================================
def build_excel(df_macro, df_guests, df_merged, df_corr):
    """Tạo file Excel đa sheet, professional formatting."""
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    _build_cover_sheet(wb, df_merged)
    _build_data_sheet(wb, df_merged)
    _build_correlation_sheet(wb, df_corr)
    _build_interest_rate_impact_sheet(wb, df_merged)
    _build_trend_charts_sheet(wb, df_merged)
    _build_insight_sheet(wb, df_corr, df_merged)
    
    wb.save(OUTPUT_FILE)
    logger.info(f"✅ Saved: {OUTPUT_FILE}")


def _apply_header_row(ws, row, cols, extra_fill=None):
    """Apply style to header row."""
    fill = extra_fill or HEADER_FILL
    for col_idx, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _auto_width(ws, min_w=10, max_w=30):
    """Auto-fit column widths."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, min_w), max_w)


def _build_cover_sheet(wb, df_merged):
    """Sheet 1: Bìa + Tổng quan."""
    ws = wb.create_sheet("📊 Tổng Quan")
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "BÁO CÁO PHÂN TÍCH TƯƠNG QUAN KINH TẾ VĨ MÔ"
    ws['A1'].font = Font(name='Calibri', bold=True, color=DARK_BLUE, size=20)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 45
    
    ws.merge_cells('A2:H2')
    ws['A2'] = "Tác động của các yếu tố vĩ mô đến lượt khách nhà hàng F&B"
    ws['A2'].font = Font(name='Calibri', color=MEDIUM_BLUE, size=13)
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 25
    
    ws.merge_cells('A3:H3')
    ws['A3'] = f"Kỳ phân tích: 01/2022 — {CURRENT_DATE.strftime('%m/%Y')}  |  Ngày tạo: {CURRENT_DATE.strftime('%d/%m/%Y')}"
    ws['A3'].font = Font(name='Calibri', color='888888', size=10, italic=True)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # --- KPI Summary ---
    row = 5
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "CHỈ SỐ HIỆN TẠI (Tháng gần nhất)"
    ws[f'A{row}'].font = Font(name='Calibri', bold=True, color=WHITE, size=12)
    ws[f'A{row}'].fill = PatternFill('solid', fgColor=ACCENT_BLUE)
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    for c in range(1, 5):
        ws.cell(row=row, column=c).fill = PatternFill('solid', fgColor=ACCENT_BLUE)
    
    if not df_merged.empty:
        latest = df_merged.iloc[-1]
        first = df_merged.iloc[0]
        
        first_label = f"{int(first['Month']):02d}/{int(first['Year'])}"
        
        kpis = [
            ("Giá Vàng SJC", f"{latest['Gia_Vang_SJC_Trieu']:.0f} triệu/lượng",
             f"{((latest['Gia_Vang_SJC_Trieu']/first['Gia_Vang_SJC_Trieu'])-1)*100:+.0f}% vs {first_label}"),
            ("Giá Xăng RON95", f"{latest['Gia_Xang_RON95_KD']:.0f} nghìn đ/lít",
             f"{((latest['Gia_Xang_RON95_KD']/first['Gia_Xang_RON95_KD'])-1)*100:+.0f}% vs {first_label}"),
            ("LS Vietcombank 12T", f"{latest['LS_Huy_Dong_12T_Pct']:.1f}%/năm",
             f"{latest['LS_Huy_Dong_12T_Pct'] - first['LS_Huy_Dong_12T_Pct']:+.1f}pp vs {first_label}"),
            ("LS Cho Vay BQ", f"{latest['LS_Cho_Vay_BQ_Pct']:.1f}%/năm",
             f"{latest['LS_Cho_Vay_BQ_Pct'] - first['LS_Cho_Vay_BQ_Pct']:+.1f}pp vs {first_label}"),
        ]
        
        for i, (label, value, change) in enumerate(kpis):
            r = row + 1 + i
            ws.cell(row=r, column=1, value=label).font = Font(name='Calibri', bold=True, size=10)
            ws.cell(row=r, column=1).fill = PatternFill('solid', fgColor=LIGHT_BLUE)
            ws.cell(row=r, column=2, value=value).font = Font(name='Calibri', size=10)
            ws.cell(row=r, column=2).alignment = DATA_ALIGN
            ws.cell(row=r, column=3, value=change).font = Font(name='Calibri', size=9, italic=True, color='666666')
            for c in range(1, 5):
                ws.cell(row=r, column=c).border = THIN_BORDER
        
        # Guest KPIs
        row_g = row + 6
        ws.merge_cells(f'A{row_g}:D{row_g}')
        ws[f'A{row_g}'] = "DỮ LIỆU KHÁCH"
        ws[f'A{row_g}'].font = Font(name='Calibri', bold=True, color=WHITE, size=12)
        ws[f'A{row_g}'].fill = PatternFill('solid', fgColor=GREEN_GOOD)
        for c in range(1, 5):
            ws.cell(row=row_g, column=c).fill = PatternFill('solid', fgColor=GREEN_GOOD)
        
        guest_kpis = [
            ("Tổng khách (kỳ phân tích)", f"{df_merged['Total_Guests'].sum():,.0f} lượt"),
            ("Tổng doanh thu (kỳ phân tích)", f"{df_merged['Total_Revenue'].sum():,.0f} VNĐ"),
            ("Số nhà hàng hoạt động", f"~{latest.get('Active_Restaurants', 'N/A')} nhà hàng"),
            ("Khách TB/ngày (tháng gần nhất)", f"{latest.get('Avg_Daily_Guests', 'N/A'):,.0f} lượt/ngày"),
            ("DT TB/khách (tháng gần nhất)", f"{latest.get('Avg_Revenue_Per_Guest', 0):,.0f} VNĐ"),
            ("Số tháng dữ liệu", f"{len(df_merged)} tháng"),
        ]
        
        for i, (label, value) in enumerate(guest_kpis):
            r = row_g + 1 + i
            ws.cell(row=r, column=1, value=label).font = Font(name='Calibri', bold=True, size=10)
            ws.cell(row=r, column=1).fill = PatternFill('solid', fgColor='E2EFDA')
            ws.cell(row=r, column=2, value=value).font = Font(name='Calibri', size=10)
            ws.cell(row=r, column=2).alignment = DATA_ALIGN
            for c in range(1, 5):
                ws.cell(row=r, column=c).border = THIN_BORDER
    
    # Set widths
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 15


def _build_data_sheet(wb, df_merged):
    """Sheet 2: Dữ liệu chi tiết theo tháng."""
    ws = wb.create_sheet("📋 Dữ Liệu Tháng")
    
    headers = [
        'Tháng/Năm', 'Năm', 'Tháng',
        'Giá Vàng SJC\n(triệu/lượng)', 'Giá Xăng RON95\n(nghìn đ/lít)',
        'LS VCB 12T\n(%/năm)', 'LS Cho Vay BQ\n(%/năm)',
        'Tổng Khách\n(lượt)', 'Tổng Doanh Thu\n(VNĐ)', 'Số Giao Dịch',
        'NH Hoạt Động', 'Ngày HĐ',
        'Khách TB/Ngày', 'Khách TB/NH/Ngày',
        'DT TB/Khách\n(VNĐ)', 'DT TB/Ngày\n(VNĐ)',
    ]
    
    _apply_header_row(ws, 1, headers)
    ws.row_dimensions[1].height = 35
    
    for idx, row_data in df_merged.iterrows():
        r = idx + 2
        ws.cell(row=r, column=1, value=row_data.get('Date_Label', '')).font = DATA_FONT
        ws.cell(row=r, column=1).alignment = DATA_ALIGN
        ws.cell(row=r, column=2, value=int(row_data['Year'])).font = DATA_FONT
        ws.cell(row=r, column=2).alignment = DATA_ALIGN
        ws.cell(row=r, column=3, value=int(row_data['Month'])).font = DATA_FONT
        ws.cell(row=r, column=3).alignment = DATA_ALIGN
        
        ws.cell(row=r, column=4, value=row_data['Gia_Vang_SJC_Trieu']).font = DATA_FONT
        ws.cell(row=r, column=4).alignment = DATA_ALIGN
        ws.cell(row=r, column=4).number_format = '#,##0.0'
        
        ws.cell(row=r, column=5, value=row_data['Gia_Xang_RON95_KD']).font = DATA_FONT
        ws.cell(row=r, column=5).alignment = DATA_ALIGN
        ws.cell(row=r, column=5).number_format = '#,##0.0'
        
        ws.cell(row=r, column=6, value=row_data['LS_Huy_Dong_12T_Pct']).font = DATA_FONT
        ws.cell(row=r, column=6).alignment = DATA_ALIGN
        ws.cell(row=r, column=6).number_format = '0.0'
        
        ws.cell(row=r, column=7, value=row_data['LS_Cho_Vay_BQ_Pct']).font = DATA_FONT
        ws.cell(row=r, column=7).alignment = DATA_ALIGN
        ws.cell(row=r, column=7).number_format = '0.0'
        
        for col_idx, field in [(8, 'Total_Guests'), (9, 'Total_Revenue'), (10, 'Total_Transactions'),
                                (11, 'Active_Restaurants'), (12, 'Active_Days')]:
            val = row_data.get(field)
            ws.cell(row=r, column=col_idx, value=int(val) if pd.notna(val) and val else '').font = DATA_FONT
            ws.cell(row=r, column=col_idx).alignment = DATA_ALIGN
            ws.cell(row=r, column=col_idx).number_format = '#,##0'
        
        for col_idx, field in [(13, 'Avg_Daily_Guests'), (14, 'Avg_Guests_Per_Restaurant')]:
            val = row_data.get(field)
            ws.cell(row=r, column=col_idx, value=round(float(val), 1) if pd.notna(val) else '').font = DATA_FONT
            ws.cell(row=r, column=col_idx).alignment = DATA_ALIGN
            ws.cell(row=r, column=col_idx).number_format = '#,##0.0'
        
        for col_idx, field in [(15, 'Avg_Revenue_Per_Guest'), (16, 'Avg_Daily_Revenue')]:
            val = row_data.get(field)
            ws.cell(row=r, column=col_idx, value=int(val) if pd.notna(val) and val else '').font = DATA_FONT
            ws.cell(row=r, column=col_idx).alignment = DATA_ALIGN
            ws.cell(row=r, column=col_idx).number_format = '#,##0'
        
        # Alternate row colors
        if idx % 2 == 1:
            for c in range(1, 17):
                ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor=LIGHT_GRAY)
        
        for c in range(1, 17):
            ws.cell(row=r, column=c).border = THIN_BORDER
    
    _auto_width(ws, min_w=12, max_w=22)
    
    # Freeze header
    ws.freeze_panes = 'A2'


def _build_correlation_sheet(wb, df_corr):
    """Sheet 3: Kết quả tương quan Pearson (Khách + Doanh Thu)."""
    ws = wb.create_sheet("📈 Tương Quan Pearson")
    
    # Title
    ws.merge_cells('A1:G1')
    ws['A1'] = "KẾT QUẢ PHÂN TÍCH TƯƠNG QUAN PEARSON"
    ws['A1'].font = Font(name='Calibri', bold=True, color=DARK_BLUE, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:G2')
    ws['A2'] = "Tương quan giữa yếu tố vĩ mô với lượt khách VÀ doanh thu nhà hàng F&B"
    ws['A2'].font = Font(name='Calibri', color='666666', size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Legend
    row_legend = 4
    ws[f'A{row_legend}'] = "THANG ĐO:"
    ws[f'A{row_legend}'].font = Font(name='Calibri', bold=True, size=10)
    
    legend_items = [
        ('B', "|r| ≥ 0.7 → Mạnh", PatternFill('solid', fgColor='C6EFCE')),
        ('C', "|r| 0.4-0.7 → Trung bình", PatternFill('solid', fgColor='FFEB9C')),
        ('D', "|r| 0.2-0.4 → Yếu", PatternFill('solid', fgColor='F4CCCC')),
        ('E', "|r| < 0.2 → Không đáng kể", PatternFill('solid', fgColor='D9D9D9')),
    ]
    for col_l, text, fill in legend_items:
        ws[f'{col_l}{row_legend}'] = text
        ws[f'{col_l}{row_legend}'].font = Font(name='Calibri', size=9)
        ws[f'{col_l}{row_legend}'].fill = fill
        ws[f'{col_l}{row_legend}'].border = THIN_BORDER
    
    # Headers
    row_h = 6
    corr_headers = [
        'Chỉ Tiêu', 'Yếu Tố Kinh Tế', 'Hệ Số r', '|r|', 'Hướng', 'Mức Độ', 'Diễn Giải'
    ]
    _apply_header_row(ws, row_h, corr_headers)
    
    # Label mapping
    label_map = {
        'Gia_Vang_SJC_Trieu': '🥇 Giá Vàng SJC',
        'Gia_Xang_RON95_KD': '⛽ Giá Xăng RON95',
        'LS_Huy_Dong_12T_Pct': '🏦 LS Vietcombank 12T',
        'LS_Cho_Vay_BQ_Pct': '💰 LS Cho Vay BQ',
    }
    
    r = row_h + 1
    prev_target = None
    for _, row_data in df_corr.iterrows():
        target = row_data.get('Target', 'Lượt Khách')
        
        # Section header when target changes
        if target != prev_target:
            if prev_target is not None:
                r += 1  # spacer
            icon = "👥" if "Khách" in target else "💵"
            ws.merge_cells(f'A{r}:G{r}')
            ws[f'A{r}'] = f"{icon}  TƯƠNG QUAN VỚI {target.upper()}"
            ws[f'A{r}'].font = Font(name='Calibri', bold=True, color=WHITE, size=11)
            section_fill = PatternFill('solid', fgColor=ACCENT_BLUE if "Khách" in target else GREEN_GOOD)
            for c in range(1, 8):
                ws.cell(row=r, column=c).fill = section_fill
            r += 1
            prev_target = target
        
        ws.cell(row=r, column=1, value=target).font = Font(name='Calibri', size=9, italic=True)
        ws.cell(row=r, column=1).alignment = DATA_ALIGN
        
        ws.cell(row=r, column=2, value=label_map.get(row_data['Yeu_To'], row_data['Yeu_To']))
        ws.cell(row=r, column=2).font = Font(name='Calibri', bold=True, size=10)
        ws.cell(row=r, column=2).alignment = DATA_ALIGN_LEFT
        
        r_val = row_data['He_So_r']
        ws.cell(row=r, column=3, value=r_val)
        ws.cell(row=r, column=3).number_format = '0.0000'
        ws.cell(row=r, column=3).alignment = DATA_ALIGN
        ws.cell(row=r, column=3).font = Font(name='Calibri', bold=True, size=11,
                                              color=GREEN_GOOD if r_val > 0 else RED_ALERT)
        
        ws.cell(row=r, column=4, value=row_data['Abs_r'])
        ws.cell(row=r, column=4).number_format = '0.0000'
        ws.cell(row=r, column=4).alignment = DATA_ALIGN
        
        ws.cell(row=r, column=5, value=row_data['Huong']).alignment = DATA_ALIGN
        ws.cell(row=r, column=6, value=row_data['Muc_Do']).alignment = DATA_ALIGN
        
        abs_r = row_data['Abs_r']
        if abs_r >= 0.7:
            strength_fill = PatternFill('solid', fgColor='C6EFCE')
        elif abs_r >= 0.4:
            strength_fill = PatternFill('solid', fgColor='FFEB9C')
        elif abs_r >= 0.2:
            strength_fill = PatternFill('solid', fgColor='F4CCCC')
        else:
            strength_fill = PatternFill('solid', fgColor='D9D9D9')
        ws.cell(row=r, column=6).fill = strength_fill
        ws.cell(row=r, column=6).font = Font(name='Calibri', bold=True, size=10)
        
        ws.cell(row=r, column=7, value=row_data['Y_Nghia'])
        ws.cell(row=r, column=7).font = Font(name='Calibri', size=9, italic=True)
        ws.cell(row=r, column=7).alignment = Alignment(wrap_text=True, vertical='center')
        
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1
    
    # Width
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 55


def _build_interest_rate_impact_sheet(wb, df_merged):
    """Sheet 4: Phân tích tác động lãi suất Vietcombank 12T lên doanh thu/khách."""
    ws = wb.create_sheet("🏦 Tác Động Lãi Suất")
    
    # Title
    ws.merge_cells('A1:I1')
    ws['A1'] = "PHÂN TÍCH TÁC ĐỘNG LÃI SUẤT VIETCOMBANK 12T LÊN DOANH THU & LƯỢT KHÁCH"
    ws['A1'].font = Font(name='Calibri', bold=True, color=DARK_BLUE, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:I2')
    ws['A2'] = ("Nguồn LS: Vietcombank kỳ hạn 12 tháng (tại quầy) | "
                f"Kỳ phân tích: 01/2022 — {CURRENT_DATE.strftime('%m/%Y')}")
    ws['A2'].font = Font(name='Calibri', color='888888', size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    if df_merged.empty or len(df_merged) < 3:
        ws['A4'] = "Không đủ dữ liệu để phân tích."
        return
    
    # --- Calculate month-over-month changes ---
    df = df_merged.copy().reset_index(drop=True)
    df['LS_Change'] = df['LS_Huy_Dong_12T_Pct'].diff()
    df['LS_Direction'] = df['LS_Change'].apply(
        lambda x: '📈 TĂNG' if pd.notna(x) and x > 0.05 else ('📉 GIẢM' if pd.notna(x) and x < -0.05 else '➡️ Ổn định')
    )
    
    has_revenue = df['Total_Revenue'].sum() > 0
    
    if has_revenue:
        df['Revenue_MoM'] = df['Total_Revenue'].pct_change() * 100
        df['Revenue_Bil'] = df['Total_Revenue'] / 1e9
    
    df['Guests_MoM'] = df['Total_Guests'].pct_change() * 100
    df['Guests_K'] = df['Total_Guests'] / 1000
    
    # ╔══════════════════════════════════════════════════════════════════════╗
    # ║  PHƯƠNG PHÁP TÍNH: Mỗi giai đoạn bắt đầu từ tháng BASELINE        ║
    # ║  (tháng cuối cùng LS còn ổn định TRƯỚC KHI bắt đầu thay đổi).    ║
    # ║  → % Thay đổi Khách/DT = so sánh tháng CUỐI giai đoạn vs BASELINE║
    # ║  → Đo đúng TÁC ĐỘNG TOÀN BỘ của chu kỳ thay đổi lãi suất.       ║
    # ╚══════════════════════════════════════════════════════════════════════╝
    rate_periods = []
    
    # Period 1: TĂNG mạnh — Baseline T08/2022 (LS=5.50%) → Cuối T03/2023 (LS=7.40%)
    # T08/2022 là tháng cuối cùng LS ổn định 5.50% trước khi tăng mạnh từ T09
    p1_start = df[(df['Year'] == 2022) & (df['Month'] == 8)]
    p1_end = df[(df['Year'] == 2023) & (df['Month'] == 3)]
    if not p1_start.empty and not p1_end.empty:
        rate_periods.append({
            'name': 'Giai đoạn TĂNG mạnh:\nT09/2022 → T03/2023',
            'baseline_label': 'T08/2022',
            'detail': 'VCB tăng từ 5.50% lên 7.40% (+1.90pp)',
            'start_idx': p1_start.index[0],
            'end_idx': p1_end.index[0],
            'direction': 'TĂNG',
        })
    
    # Period 2: GIẢM mạnh — Baseline T03/2023 (LS=7.40%) → Cuối T03/2024 (LS=4.60%)
    # T03/2023 là tháng cuối cùng LS ở đỉnh 7.40% trước khi giảm từ T04
    p2_start = df[(df['Year'] == 2023) & (df['Month'] == 3)]
    p2_end = df[(df['Year'] == 2024) & (df['Month'] == 3)]
    if not p2_start.empty and not p2_end.empty:
        rate_periods.append({
            'name': 'Giai đoạn GIẢM mạnh:\nT04/2023 → T03/2024',
            'baseline_label': 'T03/2023',
            'detail': 'VCB giảm từ 7.40% xuống 4.60% (-2.80pp)',
            'start_idx': p2_start.index[0],
            'end_idx': p2_end.index[0],
            'direction': 'GIẢM',
        })
    
    # Period 3: ỔN ĐỊNH thấp — T03/2024 (LS=4.60%) → T12/2025 (LS=4.80%)
    # Không cần baseline riêng vì LS gần như không đổi
    p3_start = df[(df['Year'] == 2024) & (df['Month'] == 3)]
    p3_end = df[(df['Year'] == 2025) & (df['Month'] == 12)]
    if not p3_start.empty and not p3_end.empty:
        rate_periods.append({
            'name': 'Giai đoạn ỔN ĐỊNH\nthấp: T03/2024 →\nT12/2025',
            'baseline_label': None,  # Không cần baseline, LS ổn định
            'detail': 'VCB duy trì 4.60-4.80% — đáy lịch sử',
            'start_idx': p3_start.index[0],
            'end_idx': p3_end.index[0],
            'direction': 'ỔN ĐỊNH',
        })
    
    # Period 4: TĂNG trở lại — Baseline T12/2025 (LS=4.80%) → Cuối T03/2026 (LS=5.90%)
    # T12/2025 là tháng cuối cùng LS ổn định 4.80% trước khi tăng từ T01/2026
    p4_start = df[(df['Year'] == 2025) & (df['Month'] == 12)]
    p4_end = df[(df['Year'] == 2026) & (df['Month'] == 3)]
    if not p4_start.empty and not p4_end.empty:
        rate_periods.append({
            'name': 'Giai đoạn TĂNG trở lại:\nT01/2026 → T03/2026',
            'baseline_label': 'T12/2025',
            'detail': 'VCB tăng từ 4.80% lên 5.90% (+1.10pp)',
            'start_idx': p4_start.index[0],
            'end_idx': p4_end.index[0],
            'direction': 'TĂNG',
        })
    
    # ============================================
    # Section 1: Summary of Rate Periods
    # ============================================
    row = 4
    ws.merge_cells(f'A{row}:I{row}')
    ws[f'A{row}'] = "1️⃣  TỔNG HỢP CÁC GIAI ĐOẠN BIẾN ĐỘNG LÃI SUẤT VCB 12T"
    ws[f'A{row}'].font = Font(name='Calibri', bold=True, color=WHITE, size=12)
    for c in range(1, 10):
        ws.cell(row=row, column=c).fill = PatternFill('solid', fgColor=ACCENT_BLUE)
    row += 1
    
    summary_headers = ['Giai Đoạn', 'Hướng', 'LS Đầu (%)', 'LS Cuối (%)', 'Thay Đổi (pp)',
                       'Khách TB/Tháng', 'Thay Đổi Khách (%)', 'DT TB/Tháng (tỷ)', 'Thay Đổi DT (%)']
    _apply_header_row(ws, row, summary_headers)
    row += 1
    
    for period in rate_periods:
        s_idx = period['start_idx']
        e_idx = period['end_idx']
        period_data = df.iloc[s_idx:e_idx + 1]
        
        if period_data.empty:
            continue
        
        # LS: lấy từ tháng baseline (đầu) đến tháng cuối
        ls_start = period_data.iloc[0]['LS_Huy_Dong_12T_Pct']
        ls_end = period_data.iloc[-1]['LS_Huy_Dong_12T_Pct']
        ls_change = ls_end - ls_start
        
        # Khách & DT: trung bình toàn giai đoạn, % thay đổi = cuối vs baseline
        avg_guests = period_data['Total_Guests'].mean()
        baseline_guests = period_data.iloc[0]['Total_Guests']  # Tháng baseline
        end_guests = period_data.iloc[-1]['Total_Guests']       # Tháng cuối
        guests_chg = ((end_guests / baseline_guests) - 1) * 100 if baseline_guests > 0 else 0
        
        if has_revenue:
            avg_rev_bil = period_data['Total_Revenue'].mean() / 1e9
            baseline_rev = period_data.iloc[0]['Total_Revenue']
            end_rev = period_data.iloc[-1]['Total_Revenue']
            rev_chg = ((end_rev / baseline_rev) - 1) * 100 if baseline_rev > 0 else 0
        else:
            avg_rev_bil = 0
            rev_chg = 0
        
        # Direction indicator
        dir_icon = '📈' if period['direction'] == 'TĂNG' else ('📉' if period['direction'] == 'GIẢM' else '➡️')
        
        ws.cell(row=row, column=1, value=period['name']).font = Font(name='Calibri', bold=True, size=9)
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='center')
        ws.cell(row=row, column=2, value=f"{dir_icon} {period['direction']}").alignment = DATA_ALIGN
        ws.cell(row=row, column=3, value=ls_start).alignment = DATA_ALIGN
        ws.cell(row=row, column=3).number_format = '0.00'
        ws.cell(row=row, column=4, value=ls_end).alignment = DATA_ALIGN
        ws.cell(row=row, column=4).number_format = '0.00'
        
        chg_cell = ws.cell(row=row, column=5, value=ls_change)
        chg_cell.number_format = '+0.00;-0.00'
        chg_cell.alignment = DATA_ALIGN
        chg_cell.font = Font(name='Calibri', bold=True, size=11,
                             color=RED_ALERT if ls_change > 0 else GREEN_GOOD)
        
        ws.cell(row=row, column=6, value=round(avg_guests)).alignment = DATA_ALIGN
        ws.cell(row=row, column=6).number_format = '#,##0'
        
        guest_chg_cell = ws.cell(row=row, column=7, value=round(guests_chg, 1))
        guest_chg_cell.number_format = '+0.0;-0.0'
        guest_chg_cell.alignment = DATA_ALIGN
        guest_chg_cell.font = Font(name='Calibri', bold=True,
                                    color=GREEN_GOOD if guests_chg > 0 else RED_ALERT)
        
        if has_revenue:
            ws.cell(row=row, column=8, value=round(avg_rev_bil, 1)).alignment = DATA_ALIGN
            ws.cell(row=row, column=8).number_format = '#,##0.0'
            
            rev_chg_cell = ws.cell(row=row, column=9, value=round(rev_chg, 1))
            rev_chg_cell.number_format = '+0.0;-0.0'
            rev_chg_cell.alignment = DATA_ALIGN
            rev_chg_cell.font = Font(name='Calibri', bold=True,
                                      color=GREEN_GOOD if rev_chg > 0 else RED_ALERT)
        
        # Background color based on direction
        if period['direction'] == 'TĂNG':
            bg = PatternFill('solid', fgColor='FCE4EC')  # light red
        elif period['direction'] == 'GIẢM':
            bg = PatternFill('solid', fgColor='E8F5E9')  # light green
        else:
            bg = PatternFill('solid', fgColor='FFF8E1')  # light yellow
        
        for c in range(1, 10):
            ws.cell(row=row, column=c).fill = bg
            ws.cell(row=row, column=c).border = THIN_BORDER
        
        ws.row_dimensions[row].height = 40
        row += 1
    
    # --- Methodology note for Section 1 ---
    row += 1
    note_lines = [
        "📌 CÁCH TÍNH:",
        "• LS Đầu / LS Cuối: Lãi suất VCB 12T tại tháng đầu tiên và cuối cùng của giai đoạn (bao gồm tháng baseline).",
        "• Khách TB/Tháng: Trung bình cộng lượt khách tất cả các tháng trong giai đoạn.",
        "• Thay Đổi Khách (%): So sánh lượt khách tháng CUỐI giai đoạn vs tháng BASELINE (tháng cuối cùng trước khi LS bắt đầu thay đổi).",
        "  Ví dụ: Giai đoạn TĂNG mạnh T09/2022→T03/2023: so sánh T03/2023 vs T08/2022 (baseline, LS còn 5.50%).",
        "• Thay Đổi DT (%): Tương tự, so sánh doanh thu tháng CUỐI vs tháng BASELINE.",
        "• Mục đích: Đo lường TÁC ĐỘNG TOÀN BỘ của chu kỳ thay đổi lãi suất lên kinh doanh nhà hàng.",
    ]
    for note in note_lines:
        ws.merge_cells(f'A{row}:I{row}')
        ws.cell(row=row, column=1, value=note)
        is_title = note.startswith("📌")
        ws.cell(row=row, column=1).font = Font(
            name='Calibri', size=10 if is_title else 9,
            bold=is_title, italic=not is_title,
            color=DARK_BLUE if is_title else '555555'
        )
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='center')
        for c in range(1, 10):
            ws.cell(row=row, column=c).fill = PatternFill('solid', fgColor='FFF8E1')
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1
    
    row += 1
    
    # ============================================
    # Section 2: Month-by-month detail table
    # ============================================
    ws.merge_cells(f'A{row}:I{row}')
    ws[f'A{row}'] = "2️⃣  CHI TIẾT THÁNG: LÃI SUẤT VCB 12T vs DOANH THU & KHÁCH"
    ws[f'A{row}'].font = Font(name='Calibri', bold=True, color=WHITE, size=12)
    for c in range(1, 10):
        ws.cell(row=row, column=c).fill = PatternFill('solid', fgColor=GREEN_GOOD)
    row += 1
    
    detail_headers = ['Tháng/Năm', 'LS VCB 12T (%)', 'Thay Đổi LS (pp)', 'Xu Hướng',
                      'Tổng Khách', 'Khách MoM (%)', 'DT (tỷ VNĐ)', 'DT MoM (%)', 'Nhận Xét']
    _apply_header_row(ws, row, detail_headers)
    row += 1
    
    for idx, r_data in df.iterrows():
        lbl = r_data.get('Date_Label', f"{int(r_data['Month']):02d}/{int(r_data['Year'])}")
        
        ws.cell(row=row, column=1, value=lbl).alignment = DATA_ALIGN
        ws.cell(row=row, column=1).font = Font(name='Calibri', bold=True, size=10)
        
        # LS VCB 12T
        ws.cell(row=row, column=2, value=r_data['LS_Huy_Dong_12T_Pct']).alignment = DATA_ALIGN
        ws.cell(row=row, column=2).number_format = '0.00'
        ws.cell(row=row, column=2).font = Font(name='Calibri', bold=True, size=10)
        
        # LS Change
        ls_chg = r_data.get('LS_Change', None)
        if pd.notna(ls_chg):
            chg_cell = ws.cell(row=row, column=3, value=round(ls_chg, 2))
            chg_cell.number_format = '+0.00;-0.00'
            chg_cell.alignment = DATA_ALIGN
            if abs(ls_chg) > 0.05:
                chg_cell.font = Font(name='Calibri', bold=True, size=10,
                                     color=RED_ALERT if ls_chg > 0 else GREEN_GOOD)
        else:
            ws.cell(row=row, column=3, value='—').alignment = DATA_ALIGN
        
        # Direction
        ws.cell(row=row, column=4, value=r_data.get('LS_Direction', '')).alignment = DATA_ALIGN
        
        # Guests
        ws.cell(row=row, column=5, value=int(r_data['Total_Guests'])).alignment = DATA_ALIGN
        ws.cell(row=row, column=5).number_format = '#,##0'
        
        # Guests MoM
        g_mom = r_data.get('Guests_MoM', None)
        if pd.notna(g_mom):
            g_cell = ws.cell(row=row, column=6, value=round(g_mom, 1))
            g_cell.number_format = '+0.0;-0.0'
            g_cell.alignment = DATA_ALIGN
            g_cell.font = Font(name='Calibri', color=GREEN_GOOD if g_mom >= 0 else RED_ALERT)
        else:
            ws.cell(row=row, column=6, value='—').alignment = DATA_ALIGN
        
        # Revenue
        if has_revenue:
            rev_bil = r_data.get('Revenue_Bil', 0)
            ws.cell(row=row, column=7, value=round(rev_bil, 1) if pd.notna(rev_bil) else 0).alignment = DATA_ALIGN
            ws.cell(row=row, column=7).number_format = '#,##0.0'
            
            r_mom = r_data.get('Revenue_MoM', None)
            if pd.notna(r_mom):
                r_cell = ws.cell(row=row, column=8, value=round(r_mom, 1))
                r_cell.number_format = '+0.0;-0.0'
                r_cell.alignment = DATA_ALIGN
                r_cell.font = Font(name='Calibri', color=GREEN_GOOD if r_mom >= 0 else RED_ALERT)
            else:
                ws.cell(row=row, column=8, value='—').alignment = DATA_ALIGN
        
        # Assessment comment
        comment = ''
        if pd.notna(ls_chg) and abs(ls_chg) > 0.2:
            if ls_chg > 0 and pd.notna(g_mom) and g_mom < 0:
                comment = '⚠️ LS tăng + Khách giảm'
            elif ls_chg > 0 and pd.notna(g_mom) and g_mom >= 0:
                comment = '✅ LS tăng nhưng Khách vẫn tăng'
            elif ls_chg < 0 and pd.notna(g_mom) and g_mom > 0:
                comment = '✅ LS giảm → Khách tăng'
            elif ls_chg < 0 and pd.notna(g_mom) and g_mom < 0:
                comment = '⚠️ LS giảm nhưng Khách vẫn giảm'
        
        ws.cell(row=row, column=9, value=comment).font = Font(name='Calibri', size=9)
        ws.cell(row=row, column=9).alignment = Alignment(wrap_text=True, vertical='center')
        
        # Highlight rate increase months
        if pd.notna(ls_chg) and ls_chg > 0.3:
            for c in range(1, 10):
                ws.cell(row=row, column=c).fill = PatternFill('solid', fgColor='FCE4EC')
        elif pd.notna(ls_chg) and ls_chg < -0.3:
            for c in range(1, 10):
                ws.cell(row=row, column=c).fill = PatternFill('solid', fgColor='E8F5E9')
        elif idx % 2 == 1:
            for c in range(1, 10):
                ws.cell(row=row, column=c).fill = PatternFill('solid', fgColor=LIGHT_GRAY)
        
        for c in range(1, 10):
            ws.cell(row=row, column=c).border = THIN_BORDER
        
        row += 1
    
    row += 1
    
    # ============================================
    # Section 3: Key Findings
    # ============================================
    ws.merge_cells(f'A{row}:I{row}')
    ws[f'A{row}'] = "3️⃣  KẾT LUẬN: TÁC ĐỘNG CỦA LÃI SUẤT VCB 12T LÊN DOANH THU"
    ws[f'A{row}'].font = Font(name='Calibri', bold=True, color=WHITE, size=12)
    for c in range(1, 10):
        ws.cell(row=row, column=c).fill = PatternFill('solid', fgColor=ORANGE_WARN)
    row += 1
    
    # Analyze correlation between rate increases and revenue changes
    rate_up_months = df[(df['LS_Change'] > 0.05) & (df['Guests_MoM'].notna())].copy()
    rate_down_months = df[(df['LS_Change'] < -0.05) & (df['Guests_MoM'].notna())].copy()
    rate_stable = df[(df['LS_Change'].abs() <= 0.05) & (df['Guests_MoM'].notna())].copy()
    
    findings = [
        f"📊 Tổng số tháng phân tích: {len(df)} tháng (01/2022 → {CURRENT_DATE.strftime('%m/%Y')})",
        f"   • Tháng LS TĂNG: {len(rate_up_months)} tháng",
        f"   • Tháng LS GIẢM: {len(rate_down_months)} tháng",
        f"   • Tháng LS ỔN ĐỊNH: {len(rate_stable)} tháng",
        "",
    ]
    
    if len(rate_up_months) > 0:
        avg_guest_chg_up = rate_up_months['Guests_MoM'].mean()
        findings.append(f"📈 Khi VCB TĂNG lãi suất ({len(rate_up_months)} tháng):")
        findings.append(f"   • Khách thay đổi TB: {avg_guest_chg_up:+.1f}% MoM")
        if has_revenue and 'Revenue_MoM' in rate_up_months.columns:
            avg_rev_chg_up = rate_up_months['Revenue_MoM'].mean()
            findings.append(f"   • Doanh thu thay đổi TB: {avg_rev_chg_up:+.1f}% MoM")
        
        # Check if guests actually decreased when rates went up
        guest_down_when_rate_up = len(rate_up_months[rate_up_months['Guests_MoM'] < 0])
        findings.append(f"   • Số tháng khách GIẢM khi LS tăng: {guest_down_when_rate_up}/{len(rate_up_months)} "
                        f"({guest_down_when_rate_up/len(rate_up_months)*100:.0f}%)")
        findings.append("")
    
    if len(rate_down_months) > 0:
        avg_guest_chg_down = rate_down_months['Guests_MoM'].mean()
        findings.append(f"📉 Khi VCB GIẢM lãi suất ({len(rate_down_months)} tháng):")
        findings.append(f"   • Khách thay đổi TB: {avg_guest_chg_down:+.1f}% MoM")
        if has_revenue and 'Revenue_MoM' in rate_down_months.columns:
            avg_rev_chg_down = rate_down_months['Revenue_MoM'].mean()
            findings.append(f"   • Doanh thu thay đổi TB: {avg_rev_chg_down:+.1f}% MoM")
        
        guest_up_when_rate_down = len(rate_down_months[rate_down_months['Guests_MoM'] > 0])
        findings.append(f"   • Số tháng khách TĂNG khi LS giảm: {guest_up_when_rate_down}/{len(rate_down_months)} "
                        f"({guest_up_when_rate_down/len(rate_down_months)*100:.0f}%)")
        findings.append("")
    
    if len(rate_stable) > 0:
        avg_guest_chg_stable = rate_stable['Guests_MoM'].mean()
        findings.append(f"➡️ Khi VCB ỔN ĐỊNH lãi suất ({len(rate_stable)} tháng):")
        findings.append(f"   • Khách thay đổi TB: {avg_guest_chg_stable:+.1f}% MoM")
        if has_revenue and 'Revenue_MoM' in rate_stable.columns:
            avg_rev_chg_stable = rate_stable['Revenue_MoM'].mean()
            findings.append(f"   • Doanh thu thay đổi TB: {avg_rev_chg_stable:+.1f}% MoM")
        findings.append("")
    
    # Overall conclusion
    findings.append("🎯 KẾT LUẬN:")
    if len(rate_up_months) > 0 and len(rate_down_months) > 0:
        avg_up = rate_up_months['Guests_MoM'].mean() if len(rate_up_months) > 0 else 0
        avg_down = rate_down_months['Guests_MoM'].mean() if len(rate_down_months) > 0 else 0
        
        if avg_up < avg_down:
            findings.append("   → Khi LS Vietcombank TĂNG, lượt khách có xu hướng GIẢM nhiều hơn so với khi LS giảm")
            findings.append("   → LS cao → người dân ưu tiên gửi tiết kiệm → giảm chi tiêu ăn uống")
        else:
            findings.append("   → Lãi suất KHÔNG phải yếu tố quyết định chính đến lượt khách")
            findings.append("   → Các yếu tố mùa vụ, lễ hội ảnh hưởng mạnh hơn lãi suất")
    
    findings.append("   → Recommendation: Khi VCB tăng LS → chuẩn bị chiến lược khuyến mại để giữ chân khách")
    
    for f_text in findings:
        ws.cell(row=row, column=1, value=f_text)
        ws.cell(row=row, column=1).font = Font(name='Calibri', size=10)
        ws.merge_cells(f'A{row}:I{row}')
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 30
    
    # Freeze
    ws.freeze_panes = f'A{len(rate_periods) + 7 + 2}'


def _build_trend_charts_sheet(wb, df_merged):
    """Sheet 4: Charts xu hướng."""
    ws = wb.create_sheet("📉 Biểu Đồ Xu Hướng")
    
    # Write data for charts (hidden area)
    chart_data_start = 1
    chart_headers = ['Tháng', 'Giá Vàng SJC', 'Giá Xăng RON95', 'LS VCB 12T', 'LS Cho Vay', 'Tổng Khách (nghìn)']
    for ci, h in enumerate(chart_headers, 1):
        ws.cell(row=chart_data_start, column=ci, value=h).font = Font(name='Calibri', bold=True, size=9)
    
    for idx, row_data in df_merged.iterrows():
        r = chart_data_start + 1 + idx
        ws.cell(row=r, column=1, value=row_data.get('Date_Label', ''))
        ws.cell(row=r, column=2, value=row_data['Gia_Vang_SJC_Trieu'])
        ws.cell(row=r, column=3, value=row_data['Gia_Xang_RON95_KD'])
        ws.cell(row=r, column=4, value=row_data['LS_Huy_Dong_12T_Pct'])
        ws.cell(row=r, column=5, value=row_data['LS_Cho_Vay_BQ_Pct'])
        guests_k = row_data.get('Total_Guests', 0)
        ws.cell(row=r, column=6, value=round(guests_k / 1000, 1) if pd.notna(guests_k) and guests_k > 0 else None)
    
    n_rows = len(df_merged)
    data_end = chart_data_start + n_rows
    
    # --- Chart 1: Giá Vàng + Khách ---
    chart1 = LineChart()
    chart1.title = "Giá Vàng SJC vs Lượt Khách"
    chart1.style = 10
    chart1.width = 28
    chart1.height = 16
    chart1.y_axis.title = "Giá Vàng (triệu VNĐ/lượng)"
    chart1.x_axis.title = "Tháng"
    
    cats = Reference(ws, min_col=1, min_row=chart_data_start+1, max_row=data_end)
    gold_data = Reference(ws, min_col=2, min_row=chart_data_start, max_row=data_end)
    guest_data = Reference(ws, min_col=6, min_row=chart_data_start, max_row=data_end)
    
    chart1.add_data(gold_data, titles_from_data=True)
    chart1.set_categories(cats)
    
    # 2nd axis for guests
    from openpyxl.chart import LineChart as LC2
    chart1_guest = LC2()
    chart1_guest.add_data(guest_data, titles_from_data=True)
    chart1_guest.y_axis.title = "Lượt Khách (nghìn)"
    chart1_guest.y_axis.axId = 200
    chart1_guest.style = 10
    
    s1 = chart1_guest.series[0]
    s1.graphicalProperties.line.dashStyle = "dash"
    
    chart1.y_axis.crosses = "min"
    chart1 += chart1_guest
    
    ws.add_chart(chart1, "H1")
    
    # --- Chart 2: Giá Xăng + Khách ---
    chart2 = LineChart()
    chart2.title = "Giá Xăng RON95 vs Lượt Khách"
    chart2.style = 10
    chart2.width = 28
    chart2.height = 16
    chart2.y_axis.title = "Giá Xăng (nghìn đ/lít)"
    
    gas_data = Reference(ws, min_col=3, min_row=chart_data_start, max_row=data_end)
    chart2.add_data(gas_data, titles_from_data=True)
    chart2.set_categories(cats)
    
    chart2_guest = LC2()
    chart2_guest.add_data(guest_data, titles_from_data=True)
    chart2_guest.y_axis.title = "Lượt Khách (nghìn)"
    chart2_guest.y_axis.axId = 200
    s2 = chart2_guest.series[0]
    s2.graphicalProperties.line.dashStyle = "dash"
    chart2 += chart2_guest
    
    ws.add_chart(chart2, "H18")
    
    # --- Chart 3: Lãi suất + Khách ---
    chart3 = LineChart()
    chart3.title = "Lãi Suất Vietcombank 12T vs Lượt Khách"
    chart3.style = 10
    chart3.width = 28
    chart3.height = 16
    chart3.y_axis.title = "Lãi Suất (%/năm)"
    
    hd_data = Reference(ws, min_col=4, min_row=chart_data_start, max_row=data_end)
    cv_data = Reference(ws, min_col=5, min_row=chart_data_start, max_row=data_end)
    chart3.add_data(hd_data, titles_from_data=True)
    chart3.add_data(cv_data, titles_from_data=True)
    chart3.set_categories(cats)
    
    chart3_guest = LC2()
    chart3_guest.add_data(guest_data, titles_from_data=True)
    chart3_guest.y_axis.title = "Lượt Khách (nghìn)"
    chart3_guest.y_axis.axId = 200
    s3 = chart3_guest.series[0]
    s3.graphicalProperties.line.dashStyle = "dash"
    chart3 += chart3_guest
    
    ws.add_chart(chart3, "H35")


def _build_insight_sheet(wb, df_corr, df_merged):
    """Sheet 5: Insights + Recommendations."""
    ws = wb.create_sheet("💡 Insights")
    
    ws.merge_cells('A1:E1')
    ws['A1'] = "PHÂN TÍCH & KHUYẾN NGHỊ CHIẾN LƯỢC"
    ws['A1'].font = Font(name='Calibri', bold=True, color=DARK_BLUE, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 35
    
    row = 3
    
    # Section 1: Key Findings
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "1️⃣  KẾT QUẢ CHÍNH"
    ws[f'A{row}'].font = Font(name='Calibri', bold=True, color=WHITE, size=12)
    ws[f'A{row}'].fill = PatternFill('solid', fgColor=ACCENT_BLUE)
    for c in range(1, 6):
        ws.cell(row=row, column=c).fill = PatternFill('solid', fgColor=ACCENT_BLUE)
    
    row += 1
    if not df_corr.empty:
        label_map = {
            'Gia_Vang_SJC_Trieu': 'Giá Vàng SJC',
            'Gia_Xang_RON95_KD': 'Giá Xăng RON95',
            'LS_Huy_Dong_12T_Pct': 'LS Vietcombank 12T',
            'LS_Cho_Vay_BQ_Pct': 'LS Cho Vay BQ',
        }
        
        findings = []
        for target_label in df_corr['Target'].unique():
            target_df = df_corr[df_corr['Target'] == target_label]
            if target_df.empty:
                continue
            strongest = target_df.iloc[0]
            findings.append(
                f"📊 {target_label.upper()}:"
            )
            findings.append(
                f"  • Yếu tố ảnh hưởng MẠNH NHẤT: {label_map.get(strongest['Yeu_To'], strongest['Yeu_To'])} "
                f"(r = {strongest['He_So_r']:.2f}, {strongest['Muc_Do']})"
            )
            findings.append(f"  • Hướng: {strongest['Huong']} → {strongest['Y_Nghia']}")
            findings.append("")
            for corr_row in target_df.itertuples():
                findings.append(
                    f"  • {label_map.get(corr_row.Yeu_To, corr_row.Yeu_To)}: "
                    f"r = {corr_row.He_So_r:+.4f} ({corr_row.Muc_Do})"
                )
            findings.append("")
    else:
        findings = ["Không đủ dữ liệu để phân tích tương quan."]
    
    for f in findings:
        ws.cell(row=row, column=1, value=f)
        ws.cell(row=row, column=1).font = Font(name='Calibri', size=10)
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
    
    row += 1
    
    # Section 2: Trend Analysis
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "2️⃣  XU HƯỚNG ĐÁNG CHÚ Ý"
    ws[f'A{row}'].font = Font(name='Calibri', bold=True, color=WHITE, size=12)
    ws[f'A{row}'].fill = PatternFill('solid', fgColor=GREEN_GOOD)
    for c in range(1, 6):
        ws.cell(row=row, column=c).fill = PatternFill('solid', fgColor=GREEN_GOOD)
    row += 1
    
    if not df_merged.empty and len(df_merged) > 1:
        first = df_merged.iloc[0]
        last = df_merged.iloc[-1]
        
        gold_chg = ((last['Gia_Vang_SJC_Trieu'] / first['Gia_Vang_SJC_Trieu']) - 1) * 100
        gas_chg = ((last['Gia_Xang_RON95_KD'] / first['Gia_Xang_RON95_KD']) - 1) * 100
        
        first_guests = first.get('Total_Guests', 0)
        last_guests = last.get('Total_Guests', 0)
        if first_guests and first_guests > 0 and pd.notna(first_guests) and pd.notna(last_guests):
            guest_chg = ((last_guests / first_guests) - 1) * 100
        else:
            guest_chg = None
        
        trends = [
            f"• Giá vàng SJC: {first['Gia_Vang_SJC_Trieu']:.0f} → {last['Gia_Vang_SJC_Trieu']:.0f} triệu ({gold_chg:+.0f}%)",
            f"• Giá xăng RON95: {first['Gia_Xang_RON95_KD']:.0f} → {last['Gia_Xang_RON95_KD']:.0f} nghìn đ ({gas_chg:+.0f}%)",
            f"• LS cho vay: {first['LS_Cho_Vay_BQ_Pct']:.1f}% → {last['LS_Cho_Vay_BQ_Pct']:.1f}% "
            f"({last['LS_Cho_Vay_BQ_Pct'] - first['LS_Cho_Vay_BQ_Pct']:+.1f}pp)",
        ]
        if guest_chg is not None:
            trends.append(f"• Lượt khách: {first_guests:,.0f} → {last_guests:,.0f} ({guest_chg:+.0f}%)")
        
        first_rev = first.get('Total_Revenue', 0) or 0
        last_rev = last.get('Total_Revenue', 0) or 0
        if first_rev > 0 and last_rev > 0:
            rev_chg = ((last_rev / first_rev) - 1) * 100
            trends.append(f"• Doanh thu: {first_rev:,.0f} → {last_rev:,.0f} VNĐ ({rev_chg:+.0f}%)")
        
        for t in trends:
            ws.cell(row=row, column=1, value=t)
            ws.cell(row=row, column=1).font = Font(name='Calibri', size=10)
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
    
    row += 1
    
    # Section 3: Recommendations
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "3️⃣  KHUYẾN NGHỊ CHIẾN LƯỢC"
    ws[f'A{row}'].font = Font(name='Calibri', bold=True, color=WHITE, size=12)
    ws[f'A{row}'].fill = PatternFill('solid', fgColor=ORANGE_WARN)
    for c in range(1, 6):
        ws.cell(row=row, column=c).fill = PatternFill('solid', fgColor=ORANGE_WARN)
    row += 1
    
    recs = [
        "📌 NGẮN HẠN (1-3 tháng):",
        "   • Khi lãi suất cho vay giảm → tăng cường marketing vì sức mua cải thiện",
        "   • Khi giá xăng tăng cao → đẩy mạnh delivery/online ordering",
        "   • Tết & lễ lớn → dự trữ nguyên liệu sớm do giá cả leo thang",
        "",
        "📌 TRUNG HẠN (3-6 tháng):",
        "   • Theo dõi chính sách tiền tệ SBV → dự đoán xu hướng tiêu dùng",
        "   • Nếu vàng tiếp tục tăng mạnh → có thể signaling bất ổn, review budget",
        "   • Tối ưu menu theo phân khúc giá phù hợp sức mua",
        "",
        "📌 DÀI HẠN (6-12 tháng):",
        "   • Integrate dữ liệu vĩ mô vào model forecast (macro features)",
        "   • Xây dựng early warning system khi macro indicators xấu đi",
        "   • Diversify revenue streams (delivery, catering, events) giảm phụ thuộc dine-in",
    ]
    
    for rec in recs:
        ws.cell(row=row, column=1, value=rec)
        ws.cell(row=row, column=1).font = Font(name='Calibri', size=10)
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
    
    row += 2
    
    # Methodology
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "📖 PHƯƠNG PHÁP"
    ws[f'A{row}'].font = Font(name='Calibri', bold=True, color=DARK_BLUE, size=11)
    row += 1
    
    methods = [
        "• Hệ số tương quan Pearson (r): Đo quan hệ tuyến tính, -1 ≤ r ≤ 1",
        "• r > 0: Tương quan thuận (cùng tăng/giảm)",
        "• r < 0: Tương quan nghịch (ngược chiều)",
        "• |r| ≥ 0.7: Mạnh | 0.4-0.7: Trung bình | 0.2-0.4: Yếu | < 0.2: Không đáng kể",
        "",
        "• Dữ liệu khách: v_fact_db_payment_hub_transactions + v_fact_db_rk_dc_transactions (dữ liệu thực từ DB)",
        "• Lãi suất huy động: Vietcombank kỳ hạn 12 tháng (tại quầy) — nguồn: vietcombank.com.vn, CafeF, DNSE",
        "• Dữ liệu vĩ mô khác: SBV (LS cho vay), SJC (vàng), Petrolimex (xăng)",
        f"• Kỳ phân tích: 01/2022 — {CURRENT_DATE.strftime('%m/%Y')} ({len(df_merged)} tháng)",
    ]
    
    for m in methods:
        ws.cell(row=row, column=1, value=m)
        ws.cell(row=row, column=1).font = Font(name='Calibri', size=9, color='666666')
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
    
    ws.column_dimensions['A'].width = 70
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15


# ==========================================
# LOAD GUEST DATA FROM REAL_GUEST_DATA.CSV (PRIORITY 1 - FULL HISTORY)
# ==========================================
def load_guest_data_from_real_csv():
    """
    Load dữ liệu khách TOÀN BỘ từ real_guest_data.csv.
    File này chứa dữ liệu daily từ 2024-01-01 đến hiện tại,
    được extract từ DB. Đây là nguồn dữ liệu đầy đủ nhất.
    """
    # Check multiple possible paths
    possible_paths = [
        PROJECT_ROOT / "forecast_system" / "scripts" / "real_guest_data.csv",
        PROJECT_ROOT / "real_guest_data.csv",
    ]
    
    real_csv = None
    for p in possible_paths:
        if p.exists():
            real_csv = p
            break
    
    if real_csv is None:
        logger.warning("real_guest_data.csv not found in any known location")
        return pd.DataFrame()
    
    logger.info(f"📂 Loading FULL HISTORY from: {real_csv}")
    df = pd.read_csv(str(real_csv))
    
    if 'guests' not in df.columns or 'date' not in df.columns:
        logger.warning("Required columns 'date'/'guests' not found in real_guest_data.csv")
        return pd.DataFrame()
    
    df['date'] = pd.to_datetime(df['date'], errors='coerce')
    df = df[df['guests'] > 0].copy()  # Filter out zero-guest days
    
    if df.empty:
        logger.warning("No valid guest data in real_guest_data.csv")
        return pd.DataFrame()
    
    df['Year'] = df['date'].dt.year
    df['Month'] = df['date'].dt.month
    
    agg_spec = {
        'Total_Guests': ('guests', 'sum'),
        'Active_Days': ('date', 'nunique'),
    }
    if 'revenue' in df.columns:
        df['revenue'] = pd.to_numeric(df['revenue'], errors='coerce').fillna(0)
        agg_spec['Total_Revenue'] = ('revenue', 'sum')
    if 'total_transactions' in df.columns:
        agg_spec['Total_Transactions'] = ('total_transactions', 'sum')
    else:
        agg_spec['Total_Transactions'] = ('guests', 'count')
    if 'active_restaurants' in df.columns:
        agg_spec['Active_Restaurants'] = ('active_restaurants', 'max')
    else:
        agg_spec['Active_Restaurants'] = ('guests', 'count')
    
    monthly = df.groupby(['Year', 'Month']).agg(**agg_spec).reset_index()
    
    monthly['Avg_Daily_Guests'] = (monthly['Total_Guests'] / monthly['Active_Days']).round(0)
    monthly['Avg_Guests_Per_Restaurant'] = (
        monthly['Total_Guests'] / monthly['Active_Restaurants'] / monthly['Active_Days']
    ).round(1)
    
    if 'Total_Revenue' not in monthly.columns:
        monthly['Total_Revenue'] = 0
    monthly['Avg_Revenue_Per_Guest'] = (monthly['Total_Revenue'] / monthly['Total_Guests'].replace(0, np.nan)).round(0)
    monthly['Avg_Daily_Revenue'] = (monthly['Total_Revenue'] / monthly['Active_Days']).round(0)
    
    # Filter to only full months (skip partials like current month or edge months)
    # Keep months with >= 25 active days (allows for months with 28 days like Feb)
    full_months = monthly[monthly['Active_Days'] >= 25].copy()
    partial = monthly[monthly['Active_Days'] < 25]
    if not partial.empty:
        logger.info(f"   ⚠️  Skipped {len(partial)} partial months (< 25 days): {list(partial[['Year','Month']].apply(lambda r: f"{int(r['Year'])}/{int(r['Month']):02d}", axis=1))}")
    
    logger.info(f"✅ Loaded {len(full_months)} full months from real_guest_data.csv ({df['date'].min().date()} → {df['date'].max().date()})")
    return full_months


# ==========================================
# LOAD GUEST DATA FROM MASTER CSV (PRIORITY 2)
# ==========================================
def load_guest_data_from_master():
    """
    Load dữ liệu Actual_Guest từ Master_Forecast_Tracking.csv.
    Chỉ có dữ liệu gần đây (2026-Q1), nên dùng làm fallback.
    """
    master_csv = PROJECT_ROOT / "Master_Forecast_Tracking.csv"
    if not master_csv.exists():
        logger.warning(f"Master CSV not found: {master_csv}")
        return pd.DataFrame()
    
    logger.info(f"📂 Loading from Master CSV: {master_csv}")
    df = pd.read_csv(str(master_csv), low_memory=False)
    
    if 'Actual_Guest' not in df.columns:
        logger.warning("Column 'Actual_Guest' not found in master CSV")
        return pd.DataFrame()
    
    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
    valid = df[df['Actual_Guest'].notna() & (df['Actual_Guest'] > 0)].copy()
    
    if valid.empty:
        logger.warning("No valid Actual_Guest data in master CSV")
        return pd.DataFrame()
    
    valid['Year'] = valid['Date'].dt.year
    valid['Month'] = valid['Date'].dt.month
    valid['date_only'] = valid['Date'].dt.date
    
    monthly = valid.groupby(['Year', 'Month']).agg(
        Total_Guests=('Actual_Guest', 'sum'),
        Total_Transactions=('Actual_Guest', 'count'),
        Active_Restaurants=('Restaurant_Code', 'nunique'),
        Active_Days=('date_only', 'nunique'),
    ).reset_index()
    
    monthly['Avg_Daily_Guests'] = (monthly['Total_Guests'] / monthly['Active_Days']).round(0)
    monthly['Avg_Guests_Per_Restaurant'] = (
        monthly['Total_Guests'] / monthly['Active_Restaurants'] / monthly['Active_Days']
    ).round(1)
    
    monthly['Total_Revenue'] = 0
    monthly['Avg_Revenue_Per_Guest'] = 0
    monthly['Avg_Daily_Revenue'] = 0
    
    logger.info(f"✅ Loaded {len(monthly)} months from Master CSV")
    return monthly


# ==========================================
# MAIN
# ==========================================
def main():
    print("=" * 60)
    print("  MACRO ECONOMIC ANALYSIS - EXCEL GENERATOR")
    print("=" * 60)
    
    # 1. Generate macro data
    print("\n📊 Generating macro economic data (Vietcombank 12T rates)...")
    df_macro = generate_macro_data()
    print(f"   → {len(df_macro)} months of macro data")
    
    # ==========================================
    # 2. Load guest + revenue data (hybrid strategy)
    #    - 2022: Sales_by_Store.xlsx (Sales Act = DT, TC = Khách)
    #    - 2023+: DB → real_guest_data.csv → Master CSV
    # ==========================================
    print("\n📂 Loading guest + revenue data...")
    data_sources = []
    
    # --- Step A: Load 2022 from Sales_by_Store.xlsx ---
    print("\n   [A] Loading 2022 data from Sales_by_Store.xlsx (Sales Act / TC)...")
    df_2022 = load_guest_data_from_sales_by_store()
    if not df_2022.empty:
        # Keep only 2022 data from this source
        df_2022 = df_2022[df_2022['Year'] == 2022].copy()
        print(f"   → 2022: {len(df_2022)} months from Sales_by_Store.xlsx")
        data_sources.append("Sales_by_Store.xlsx (2022)")
    else:
        print("   ⚠️  Sales_by_Store.xlsx unavailable for 2022 data")
    
    # --- Step B: Load 2023+ from DB (primary) ---
    print("\n   [B] Loading 2023+ data from Database...")
    df_db = load_guest_data_from_db()
    
    if not df_db.empty:
        # Keep only 2023+ from DB (2022 comes from Sales_by_Store)
        df_db_filtered = df_db[df_db['Year'] >= 2023].copy()
        print(f"   → 2023+: {len(df_db_filtered)} months from Database")
        data_sources.append(f"Database (2023-{CURRENT_DATE.year})")
    else:
        df_db_filtered = pd.DataFrame()
        print("   ⚠️  DB unavailable. Trying fallback sources for 2023+ data...")
        
        # Fallback: real_guest_data.csv
        df_fallback = load_guest_data_from_real_csv()
        if not df_fallback.empty:
            df_db_filtered = df_fallback[df_fallback['Year'] >= 2023].copy()
            data_sources.append("real_guest_data.csv (2023+)")
        else:
            # Fallback: Master CSV
            df_fallback = load_guest_data_from_master()
            if not df_fallback.empty:
                df_db_filtered = df_fallback[df_fallback['Year'] >= 2023].copy()
                data_sources.append("Master_Forecast_Tracking.csv (2023+)")
    
    # --- Step C: Combine 2022 + 2023+ ---
    print("\n   [C] Combining data sources...")
    parts = []
    if not df_2022.empty:
        parts.append(df_2022)
    if not df_db_filtered.empty:
        parts.append(df_db_filtered)
    
    if not parts:
        print("❌ No data sources available. Cannot proceed.")
        print("   Please ensure Sales_by_Store.xlsx AND (DB or CSV) are accessible.")
        return
    
    df_guests = pd.concat(parts, ignore_index=True)
    
    # Remove duplicates (prefer first source = Sales_by_Store for 2022)
    df_guests = df_guests.drop_duplicates(subset=['Year', 'Month'], keep='first')
    df_guests = df_guests.sort_values(['Year', 'Month']).reset_index(drop=True)
    
    data_source = " + ".join(data_sources)
    
    print(f"   → Combined: {len(df_guests)} months total")
    print(f"   → Sources: {data_source}")
    
    has_revenue = df_guests['Total_Revenue'].sum() > 0
    
    # Show monthly breakdown
    print("\n   MONTHLY DATA:")
    for _, row in df_guests.iterrows():
        rev_str = f"  |  💵 {int(row.get('Total_Revenue', 0)):>15,} VNĐ" if has_revenue else ""
        src_tag = "📋" if row['Year'] == 2022 else "🗄️"
        print(f"   {src_tag} {int(row['Year'])}/{int(row['Month']):02d}: 👥 {int(row['Total_Guests']):>12,} khách{rev_str}")
    
    # 3. Merge (only months that have BOTH macro + guest data)
    print("\n🔗 Merging macro + guest data...")
    df_merged = pd.merge(df_macro, df_guests, on=['Year', 'Month'], how='inner')
    print(f"   → {len(df_merged)} months merged (only months with real guest data)")
    
    if df_merged.empty:
        print("❌ No overlapping data. Check date ranges.")
        return
    
    # 4. Calculate correlations
    print("\n📈 Calculating Pearson correlations...")
    if len(df_merged) < 3:
        print(f"   ⚠️  Only {len(df_merged)} months — insufficient for reliable correlation (need ≥ 3)")
        print("   Skipping correlation analysis, generating data report only.")
        df_corr = pd.DataFrame()
    else:
        df_corr = calculate_correlations(df_merged)
        if not df_corr.empty:
            print("\n   RESULTS:")
            for _, r in df_corr.iterrows():
                label_map = {
                    'Gia_Vang_SJC_Trieu': 'Giá Vàng',
                    'Gia_Xang_RON95_KD': 'Giá Xăng',
                    'LS_Huy_Dong_12T_Pct': 'LS VCB 12T',
                    'LS_Cho_Vay_BQ_Pct': 'LS Cho Vay',
                }
                target_short = "👥" if "Khách" in r.get('Target', '') else "💵"
                print(f"   {target_short} {label_map.get(r['Yeu_To'], r['Yeu_To']):15s}  r = {r['He_So_r']:+.4f}  ({r['Muc_Do']})")
        else:
            print("   ⚠️  Not enough data points for correlation (need ≥ 5)")
    
    # 5. Build Excel
    print(f"\n📝 Building Excel report...")
    build_excel(df_macro, df_guests, df_merged, df_corr)
    
    print(f"\n{'='*60}")
    print(f"  ✅ DONE! File saved: {OUTPUT_FILE}")
    print(f"  📊 Sheets: Tổng Quan | Dữ Liệu Tháng | Tương Quan Pearson | Tác Động LS | Biểu Đồ | Insights")
    print(f"  📋 Data source: {data_source}")
    print(f"  📅 Months with real data: {len(df_merged)}")
    print(f"{'='*60}")


if __name__ == '__main__':
    main()

