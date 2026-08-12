"""
===============================================
MODEL PERFORMANCE EVALUATION - EXCEL REPORT
===============================================
Đánh giá hiệu suất Model theo:
  1. Ca làm việc (Sáng / Tối)
  2. Ngày trong tuần vs Cuối tuần
  3. Tỷ lệ forecast chính xác (Hit Rate, MAPE, MAE)
  4. Xu hướng accuracy theo thời gian

Output: Model_Performance_Report.xlsx (đa sheet, charts, conditional formatting)

Usage:
    python -m forecast_system.scripts.generate_model_performance_report
"""

import sys, os, datetime, traceback
import numpy as np
import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

from forecast_system.config.settings import PROJECT_ROOT, CURRENT_DATE, MASTER_FILE_NAME, MONITORING_CONFIG
from forecast_system.utils.logger import get_logger

logger = get_logger('model_performance')
OUTPUT_FILE = str(PROJECT_ROOT / "Model_Performance_Report.xlsx")
WEB_DATA_FILE = str(PROJECT_ROOT / "forecast_system" / "scripts" / "macro_analysis" / "performance_data.js")

# ==========================================
# STYLES
# ==========================================
DARK_BG = "1A1A2E"
BLUE_ACCENT = "4361EE"
GREEN_GOOD = "2ECC71"
RED_BAD = "E74C3C"
ORANGE_WARN = "F39C12"
PURPLE = "7B2D8E"
MORNING_COLOR = "FF9F43"
EVENING_COLOR = "6C5CE7"
WHITE = "FFFFFF"
LIGHT_GRAY = "F8F9FA"

THIN_BORDER = Border(
    left=Side(style='thin', color='DEE2E6'),
    right=Side(style='thin', color='DEE2E6'),
    top=Side(style='thin', color='DEE2E6'),
    bottom=Side(style='thin', color='DEE2E6'),
)
HEADER_FONT = Font(name='Calibri', bold=True, color=WHITE, size=11)
HEADER_FILL = PatternFill('solid', fgColor=DARK_BG)
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
DATA_FONT = Font(name='Calibri', size=10)
DATA_ALIGN = Alignment(horizontal='center', vertical='center')

HIT_THRESHOLD = MONITORING_CONFIG.get('hit_rate_threshold_abs', 15)


# ==========================================
# DATA LOADING & PREPARATION
# ==========================================
def load_master_data():
    """Load Master Forecast file and prepare for analysis."""
    logger.info(f"📂 Loading Master File: {MASTER_FILE_NAME}")
    
    if not os.path.exists(MASTER_FILE_NAME):
        logger.error(f"Master file not found: {MASTER_FILE_NAME}")
        return pd.DataFrame()
    
    df = pd.read_excel(MASTER_FILE_NAME, engine='openpyxl')
    logger.info(f"   Loaded {len(df):,} rows")
    
    # Normalize types
    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
    df = df.dropna(subset=['Date'])
    df['Final_Predicted_Guests'] = pd.to_numeric(df['Final_Predicted_Guests'], errors='coerce')
    df['Actual_Guest'] = pd.to_numeric(df['Actual_Guest'], errors='coerce')
    
    # Filter: only rows with both predicted AND actual
    mask = (
        df['Final_Predicted_Guests'].notna() &
        df['Actual_Guest'].notna() &
        (df['Final_Predicted_Guests'] >= 0) &
        (df['Actual_Guest'] >= 0)
    )
    df_valid = df[mask].copy()
    
    if df_valid.empty:
        logger.warning("No valid prediction/actual pairs found!")
        return pd.DataFrame()
    
    # Derived columns
    df_valid['Error'] = df_valid['Final_Predicted_Guests'] - df_valid['Actual_Guest']
    df_valid['Abs_Error'] = df_valid['Error'].abs()
    nz = df_valid['Actual_Guest'] > 0
    df_valid.loc[nz, 'Pct_Error'] = (df_valid.loc[nz, 'Abs_Error'] / df_valid.loc[nz, 'Actual_Guest']) * 100
    df_valid['Hit'] = (df_valid['Abs_Error'] <= HIT_THRESHOLD).astype(int)
    
    # Weekday info
    df_valid['Weekday_Name'] = df_valid['Date'].dt.day_name()
    df_valid['Day_Num'] = df_valid['Date'].dt.dayofweek  # 0=Mon, 6=Sun
    df_valid['Is_Weekend'] = df_valid['Day_Num'].isin([5, 6])
    df_valid['Day_Type'] = df_valid['Is_Weekend'].map({True: 'Cuối tuần', False: 'Ngày thường'})
    
    # Shift info
    if 'Shift' not in df_valid.columns:
        df_valid['Shift'] = 'UNKNOWN'
    df_valid['Shift'] = df_valid['Shift'].fillna('UNKNOWN')
    df_valid['Shift_Label'] = df_valid['Shift'].map({
        'MORNING': 'Ca Sáng (8h-15h30)',
        'EVENING': 'Ca Tối (15h30-23h)',
    }).fillna('Khác')
    
    df_valid['Date_Only'] = df_valid['Date'].dt.date
    df_valid['Week'] = df_valid['Date'].dt.isocalendar().week.astype(int)
    df_valid['YearMonth'] = df_valid['Date'].dt.to_period('M').astype(str)
    
    logger.info(f"   Valid pairs: {len(df_valid):,} rows")
    logger.info(f"   Date range: {df_valid['Date'].min().date()} → {df_valid['Date'].max().date()}")
    logger.info(f"   Shifts: {df_valid['Shift'].value_counts().to_dict()}")
    return df_valid


def calc_metrics(df):
    """Calculate standard metrics from a DataFrame subset."""
    if df.empty:
        return {'MAE': None, 'MAPE': None, 'WMAPE': None, 'RMSE': None, 'Bias': None, 'Hit_Rate': None, 'N': 0}
    
    mae = df['Abs_Error'].mean()
    rmse = np.sqrt((df['Error'] ** 2).mean())
    bias = df['Error'].mean()
    hit_rate = df['Hit'].mean() * 100
    n = len(df)
    
    nz = df[df['Actual_Guest'] > 0]
    mape = nz['Pct_Error'].mean() if not nz.empty else None
    total_actual = nz['Actual_Guest'].sum()
    wmape = (nz['Abs_Error'].sum() / total_actual * 100) if total_actual > 0 else None
    
    return {
        'MAE': round(mae, 2), 'MAPE': round(mape, 1) if mape else None,
        'WMAPE': round(wmape, 1) if wmape else None,
        'RMSE': round(rmse, 2), 'Bias': round(bias, 2),
        'Hit_Rate': round(hit_rate, 1), 'N': n,
    }


# ==========================================
# ANALYSIS FUNCTIONS
# ==========================================
def analyze_by_shift(df):
    """Accuracy breakdown by shift."""
    results = []
    for shift in ['MORNING', 'EVENING']:
        sub = df[df['Shift'] == shift]
        if sub.empty:
            continue
        m = calc_metrics(sub)
        m['Shift'] = shift
        m['Shift_Label'] = 'Ca Sáng' if shift == 'MORNING' else 'Ca Tối'
        results.append(m)
    return pd.DataFrame(results)


def analyze_by_shift_daytype(df):
    """Cross-analysis: Shift × Day Type (weekday/weekend)."""
    results = []
    for shift in ['MORNING', 'EVENING']:
        for is_wknd, label in [(False, 'Ngày thường'), (True, 'Cuối tuần')]:
            sub = df[(df['Shift'] == shift) & (df['Is_Weekend'] == is_wknd)]
            if sub.empty:
                continue
            m = calc_metrics(sub)
            m['Shift'] = 'Ca Sáng' if shift == 'MORNING' else 'Ca Tối'
            m['Day_Type'] = label
            m['Category'] = f"{'Ca Sáng' if shift == 'MORNING' else 'Ca Tối'} - {label}"
            results.append(m)
    return pd.DataFrame(results)


def analyze_by_weekday(df):
    """Accuracy per day of week."""
    day_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    vn_names = {'Monday': 'Thứ Hai', 'Tuesday': 'Thứ Ba', 'Wednesday': 'Thứ Tư',
                'Thursday': 'Thứ Năm', 'Friday': 'Thứ Sáu', 'Saturday': 'Thứ Bảy', 'Sunday': 'Chủ Nhật'}
    results = []
    for day in day_order:
        sub = df[df['Weekday_Name'] == day]
        if sub.empty:
            continue
        m = calc_metrics(sub)
        m['Weekday'] = day
        m['Weekday_VN'] = vn_names.get(day, day)
        m['Is_Weekend'] = day in ('Saturday', 'Sunday')
        results.append(m)
    return pd.DataFrame(results)


def analyze_by_weekday_shift(df):
    """Full matrix: weekday × shift."""
    day_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    vn_names = {'Monday': 'T2', 'Tuesday': 'T3', 'Wednesday': 'T4',
                'Thursday': 'T5', 'Friday': 'T6', 'Saturday': 'T7', 'Sunday': 'CN'}
    results = []
    for day in day_order:
        for shift in ['MORNING', 'EVENING']:
            sub = df[(df['Weekday_Name'] == day) & (df['Shift'] == shift)]
            if sub.empty:
                continue
            m = calc_metrics(sub)
            m['Weekday'] = vn_names.get(day, day)
            m['Shift'] = 'Sáng' if shift == 'MORNING' else 'Tối'
            results.append(m)
    return pd.DataFrame(results)


def analyze_weekly_trend(df):
    """Accuracy trend by week."""
    df_sorted = df.sort_values('Date')
    results = []
    for (yr, wk), grp in df_sorted.groupby([df_sorted['Date'].dt.year, 'Week']):
        m = calc_metrics(grp)
        m['Year'] = yr
        m['Week'] = wk
        dates = grp['Date_Only'].unique()
        m['Period'] = f"{min(dates)} → {max(dates)}"
        m['Week_Label'] = f"W{wk}/{yr}"
        results.append(m)
    return pd.DataFrame(results)


def analyze_accuracy_distribution(df):
    """Distribution of error ranges."""
    bins = [0, 5, 10, 15, 20, 30, 50, float('inf')]
    labels = ['≤5', '6-10', '11-15', '16-20', '21-30', '31-50', '>50']
    df_temp = df.copy()
    df_temp['Error_Bin'] = pd.cut(df_temp['Abs_Error'], bins=bins, labels=labels, right=True)
    
    dist = df_temp['Error_Bin'].value_counts().reindex(labels).fillna(0)
    total = len(df_temp)
    results = []
    cumulative = 0
    for label in labels:
        count = int(dist.get(label, 0))
        cumulative += count
        results.append({
            'Error_Range': f"{label} khách",
            'Count': count, 'Pct': round(count / total * 100, 1) if total > 0 else 0,
            'Cumulative_Pct': round(cumulative / total * 100, 1) if total > 0 else 0,
        })
    return pd.DataFrame(results)


# ==========================================
# EXCEL BUILDER
# ==========================================
def _header_row(ws, row, cols, fill=None):
    f = fill or HEADER_FILL
    for ci, val in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=val)
        c.font = HEADER_FONT; c.fill = f; c.alignment = HEADER_ALIGN; c.border = THIN_BORDER

def _auto_width(ws, min_w=10, max_w=30):
    for col_cells in ws.columns:
        mx = max((len(str(c.value or '')) for c in col_cells), default=0)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(mx + 3, min_w), max_w)

def _color_cell(ws, row, col, value, good_low=True):
    """Color cell based on value (green=good, red=bad)."""
    cell = ws.cell(row=row, column=col)
    if value is None:
        return
    if good_low:
        if value <= 10: cell.fill = PatternFill('solid', fgColor='C6EFCE')
        elif value <= 20: cell.fill = PatternFill('solid', fgColor='FFEB9C')
        elif value <= 30: cell.fill = PatternFill('solid', fgColor='FFC7CE')
        else: cell.fill = PatternFill('solid', fgColor='FF9999')
    else:
        if value >= 80: cell.fill = PatternFill('solid', fgColor='C6EFCE')
        elif value >= 60: cell.fill = PatternFill('solid', fgColor='FFEB9C')
        elif value >= 40: cell.fill = PatternFill('solid', fgColor='FFC7CE')
        else: cell.fill = PatternFill('solid', fgColor='FF9999')


def build_excel(df):
    """Build comprehensive Excel report."""
    wb = Workbook()
    wb.remove(wb.active)
    
    overall = calc_metrics(df)
    df_shift = analyze_by_shift(df)
    df_shift_day = analyze_by_shift_daytype(df)
    df_weekday = analyze_by_weekday(df)
    df_wd_shift = analyze_by_weekday_shift(df)
    df_weekly = analyze_weekly_trend(df)
    df_dist = analyze_accuracy_distribution(df)
    
    # --- SHEET 1: Dashboard ---
    ws = wb.create_sheet("📊 Dashboard")
    ws.merge_cells('A1:H1')
    ws['A1'] = "ĐÁNH GIÁ HIỆU SUẤT MODEL FORECAST"
    ws['A1'].font = Font(name='Calibri', bold=True, color=DARK_BG, size=20)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 45
    
    ws.merge_cells('A2:H2')
    ws['A2'] = f"Kỳ phân tích: {df['Date'].min().strftime('%d/%m/%Y')} — {df['Date'].max().strftime('%d/%m/%Y')}  |  Ngày tạo: {CURRENT_DATE.strftime('%d/%m/%Y')}"
    ws['A2'].font = Font(name='Calibri', color='888888', size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # KPI Cards
    r = 4
    kpi_header = PatternFill('solid', fgColor=BLUE_ACCENT)
    ws.merge_cells(f'A{r}:D{r}')
    ws[f'A{r}'] = "📋 TỔNG QUAN HIỆU SUẤT"
    ws[f'A{r}'].font = Font(name='Calibri', bold=True, color=WHITE, size=12)
    for c in range(1,5): ws.cell(row=r, column=c).fill = kpi_header
    
    kpis = [
        ("Hit Rate (|error| ≤ 15 khách)", f"{overall['Hit_Rate']}%", "Tỷ lệ dự đoán chính xác"),
        ("MAE (Sai số trung bình)", f"{overall['MAE']} khách", "Trung bình sai lệch mỗi ca"),
        ("MAPE (Sai số %)", f"{overall['MAPE']}%", "Sai số theo tỷ lệ phần trăm"),
        ("WMAPE (Sai số % có trọng số)", f"{overall['WMAPE']}%", "Weighted - ít bị ảnh hưởng bởi ca thấp điểm"),
        ("Bias (Xu hướng sai)", f"{overall['Bias']:+.1f} khách", "Dương = dự đoán thừa, Âm = dự đoán thiếu"),
        ("Số mẫu phân tích", f"{overall['N']:,}", "Tổng số ca có dữ liệu thực tế"),
    ]
    for i, (label, value, desc) in enumerate(kpis):
        row = r + 1 + i
        ws.cell(row=row, column=1, value=label).font = Font(name='Calibri', bold=True, size=10)
        ws.cell(row=row, column=1).fill = PatternFill('solid', fgColor='E8EAF6')
        ws.cell(row=row, column=2, value=value).font = Font(name='Calibri', bold=True, size=11)
        ws.cell(row=row, column=2).alignment = DATA_ALIGN
        ws.cell(row=row, column=3, value=desc).font = Font(name='Calibri', size=9, italic=True, color='666666')
        for c in range(1,5): ws.cell(row=row, column=c).border = THIN_BORDER
    
    # Shift comparison section
    r2 = r + len(kpis) + 2
    ws.merge_cells(f'A{r2}:D{r2}')
    ws[f'A{r2}'] = "🌅🌙 SO SÁNH CA SÁNG vs CA TỐI"
    ws[f'A{r2}'].font = Font(name='Calibri', bold=True, color=WHITE, size=12)
    for c in range(1,5): ws.cell(row=r2, column=c).fill = PatternFill('solid', fgColor=PURPLE)
    
    if not df_shift.empty:
        _header_row(ws, r2+1, ['Ca', 'Hit Rate %', 'MAE', 'MAPE %', 'Bias', 'N'], PatternFill('solid', fgColor='37474F'))
        for i, (_, row_data) in enumerate(df_shift.iterrows()):
            rr = r2 + 2 + i
            shift_fill = PatternFill('solid', fgColor='FFF3E0') if row_data['Shift'] == 'MORNING' else PatternFill('solid', fgColor='EDE7F6')
            ws.cell(row=rr, column=1, value=row_data['Shift_Label']).font = Font(name='Calibri', bold=True, size=10)
            ws.cell(row=rr, column=1).fill = shift_fill
            for ci, field in [(2, 'Hit_Rate'), (3, 'MAE'), (4, 'MAPE'), (5, 'Bias'), (6, 'N')]:
                val = row_data.get(field)
                ws.cell(row=rr, column=ci, value=val).font = DATA_FONT
                ws.cell(row=rr, column=ci).alignment = DATA_ALIGN
                ws.cell(row=rr, column=ci).border = THIN_BORDER
            _color_cell(ws, rr, 2, row_data.get('Hit_Rate'), good_low=False)
    
    # Weekday vs Weekend
    r3 = r2 + (len(df_shift) if not df_shift.empty else 0) + 4
    ws.merge_cells(f'A{r3}:D{r3}')
    ws[f'A{r3}'] = "📅 NGÀY THƯỜNG vs CUỐI TUẦN"
    ws[f'A{r3}'].font = Font(name='Calibri', bold=True, color=WHITE, size=12)
    for c in range(1,5): ws.cell(row=r3, column=c).fill = PatternFill('solid', fgColor='1565C0')
    
    if not df_shift_day.empty:
        _header_row(ws, r3+1, ['Phân loại', 'Hit Rate %', 'MAE', 'MAPE %', 'Bias', 'N'], PatternFill('solid', fgColor='37474F'))
        for i, (_, rd) in enumerate(df_shift_day.iterrows()):
            rr = r3 + 2 + i
            ws.cell(row=rr, column=1, value=rd['Category']).font = Font(name='Calibri', bold=True, size=10)
            for ci, field in [(2, 'Hit_Rate'), (3, 'MAE'), (4, 'MAPE'), (5, 'Bias'), (6, 'N')]:
                ws.cell(row=rr, column=ci, value=rd.get(field)).font = DATA_FONT
                ws.cell(row=rr, column=ci).alignment = DATA_ALIGN
                ws.cell(row=rr, column=ci).border = THIN_BORDER
            _color_cell(ws, rr, 2, rd.get('Hit_Rate'), good_low=False)
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    
    # --- SHEET 2: Weekday Detail ---
    ws2 = wb.create_sheet("📅 Theo Ngày Tuần")
    _header_row(ws2, 1, ['Thứ', 'Loại', 'Hit Rate %', 'MAE', 'MAPE %', 'WMAPE %', 'RMSE', 'Bias', 'N'])
    if not df_weekday.empty:
        for i, (_, rd) in enumerate(df_weekday.iterrows()):
            rr = 2 + i
            ws2.cell(row=rr, column=1, value=rd['Weekday_VN']).font = Font(name='Calibri', bold=True, size=10)
            ws2.cell(row=rr, column=2, value='Cuối tuần' if rd['Is_Weekend'] else 'Ngày thường').font = DATA_FONT
            for ci, field in [(3,'Hit_Rate'),(4,'MAE'),(5,'MAPE'),(6,'WMAPE'),(7,'RMSE'),(8,'Bias'),(9,'N')]:
                ws2.cell(row=rr, column=ci, value=rd.get(field)).font = DATA_FONT
                ws2.cell(row=rr, column=ci).alignment = DATA_ALIGN
                ws2.cell(row=rr, column=ci).border = THIN_BORDER
            bg = PatternFill('solid', fgColor='FFF9C4') if rd['Is_Weekend'] else PatternFill('solid', fgColor=LIGHT_GRAY)
            for c in range(1, 10): ws2.cell(row=rr, column=c).fill = bg if i % 2 == 0 else PatternFill()
            _color_cell(ws2, rr, 3, rd.get('Hit_Rate'), good_low=False)
            _color_cell(ws2, rr, 5, rd.get('MAPE'), good_low=True)
    _auto_width(ws2)
    
    # --- SHEET 3: Weekday × Shift Matrix ---
    ws3 = wb.create_sheet("🔀 Ngày × Ca")
    _header_row(ws3, 1, ['Thứ', 'Ca', 'Hit Rate %', 'MAE', 'MAPE %', 'Bias', 'N'])
    if not df_wd_shift.empty:
        for i, (_, rd) in enumerate(df_wd_shift.iterrows()):
            rr = 2 + i
            ws3.cell(row=rr, column=1, value=rd['Weekday']).font = Font(name='Calibri', bold=True)
            shift_fill = PatternFill('solid', fgColor='FFF3E0') if rd['Shift'] == 'Sáng' else PatternFill('solid', fgColor='EDE7F6')
            ws3.cell(row=rr, column=2, value=rd['Shift']).font = DATA_FONT
            ws3.cell(row=rr, column=2).fill = shift_fill
            for ci, field in [(3,'Hit_Rate'),(4,'MAE'),(5,'MAPE'),(6,'Bias'),(7,'N')]:
                ws3.cell(row=rr, column=ci, value=rd.get(field)).font = DATA_FONT
                ws3.cell(row=rr, column=ci).alignment = DATA_ALIGN
                ws3.cell(row=rr, column=ci).border = THIN_BORDER
            _color_cell(ws3, rr, 3, rd.get('Hit_Rate'), good_low=False)
    _auto_width(ws3)
    
    # --- SHEET 4: Weekly Trend ---
    ws4 = wb.create_sheet("📈 Xu Hướng Tuần")
    _header_row(ws4, 1, ['Tuần', 'Kỳ', 'Hit Rate %', 'MAE', 'MAPE %', 'WMAPE %', 'Bias', 'N'])
    if not df_weekly.empty:
        for i, (_, rd) in enumerate(df_weekly.iterrows()):
            rr = 2 + i
            ws4.cell(row=rr, column=1, value=rd.get('Week_Label', '')).font = DATA_FONT
            ws4.cell(row=rr, column=2, value=rd.get('Period', '')).font = Font(name='Calibri', size=9)
            for ci, field in [(3,'Hit_Rate'),(4,'MAE'),(5,'MAPE'),(6,'WMAPE'),(7,'Bias'),(8,'N')]:
                ws4.cell(row=rr, column=ci, value=rd.get(field)).font = DATA_FONT
                ws4.cell(row=rr, column=ci).alignment = DATA_ALIGN
                ws4.cell(row=rr, column=ci).border = THIN_BORDER
            _color_cell(ws4, rr, 3, rd.get('Hit_Rate'), good_low=False)
            if i % 2 == 1:
                for c in range(1,9): ws4.cell(row=rr, column=c).fill = PatternFill('solid', fgColor=LIGHT_GRAY)
    _auto_width(ws4, max_w=35)
    
    # Add chart to weekly trend
    if not df_weekly.empty and len(df_weekly) >= 2:
        chart = LineChart()
        chart.title = "Hit Rate % theo Tuần"
        chart.style = 10; chart.width = 28; chart.height = 14
        chart.y_axis.title = "Hit Rate %"; chart.y_axis.scaling.min = 0; chart.y_axis.scaling.max = 100
        data_ref = Reference(ws4, min_col=3, min_row=1, max_row=1+len(df_weekly))
        cats = Reference(ws4, min_col=1, min_row=2, max_row=1+len(df_weekly))
        chart.add_data(data_ref, titles_from_data=True); chart.set_categories(cats)
        ws4.add_chart(chart, f"J2")
    
    # --- SHEET 5: Error Distribution ---
    ws5 = wb.create_sheet("📊 Phân Bố Sai Số")
    _header_row(ws5, 1, ['Khoảng Sai Số', 'Số lượng', 'Tỷ lệ %', 'Tích lũy %'])
    if not df_dist.empty:
        for i, (_, rd) in enumerate(df_dist.iterrows()):
            rr = 2 + i
            ws5.cell(row=rr, column=1, value=rd['Error_Range']).font = Font(name='Calibri', bold=True)
            ws5.cell(row=rr, column=2, value=rd['Count']).font = DATA_FONT
            ws5.cell(row=rr, column=2).alignment = DATA_ALIGN; ws5.cell(row=rr, column=2).number_format = '#,##0'
            ws5.cell(row=rr, column=3, value=rd['Pct']).font = DATA_FONT
            ws5.cell(row=rr, column=3).alignment = DATA_ALIGN
            ws5.cell(row=rr, column=4, value=rd['Cumulative_Pct']).font = DATA_FONT
            ws5.cell(row=rr, column=4).alignment = DATA_ALIGN
            for c in range(1,5): ws5.cell(row=rr, column=c).border = THIN_BORDER
    _auto_width(ws5)
    
    # Add bar chart
    if not df_dist.empty:
        chart2 = BarChart()
        chart2.title = "Phân Bố Sai Số (Số khách)"
        chart2.style = 10; chart2.width = 24; chart2.height = 14
        data_ref = Reference(ws5, min_col=3, min_row=1, max_row=1+len(df_dist))
        cats = Reference(ws5, min_col=1, min_row=2, max_row=1+len(df_dist))
        chart2.add_data(data_ref, titles_from_data=True); chart2.set_categories(cats)
        ws5.add_chart(chart2, "F2")
    
    wb.save(OUTPUT_FILE)
    logger.info(f"✅ Excel saved: {OUTPUT_FILE}")


# ==========================================
# WEB DATA EXPORT (JS for dashboard)
# ==========================================
def export_web_data(df):
    """Export analysis results as JS data for web dashboard."""
    overall = calc_metrics(df)
    df_shift = analyze_by_shift(df)
    df_shift_day = analyze_by_shift_daytype(df)
    df_weekday = analyze_by_weekday(df)
    df_wd_shift = analyze_by_weekday_shift(df)
    df_weekly = analyze_weekly_trend(df)
    df_dist = analyze_accuracy_distribution(df)
    
    lines = [
        "/**",
        " * Model Performance Data - Auto-generated",
        f" * Generated: {CURRENT_DATE.strftime('%Y-%m-%d')}",
        f" * Period: {df['Date'].min().strftime('%Y-%m-%d')} → {df['Date'].max().strftime('%Y-%m-%d')}",
        " */",
        "",
        f"const PERF_OVERALL = {_dict_to_js(overall)};",
        "",
        f"const PERF_BY_SHIFT = {_df_to_js(df_shift)};",
        "",
        f"const PERF_SHIFT_DAYTYPE = {_df_to_js(df_shift_day)};",
        "",
        f"const PERF_BY_WEEKDAY = {_df_to_js(df_weekday)};",
        "",
        f"const PERF_WEEKDAY_SHIFT = {_df_to_js(df_wd_shift)};",
        "",
        f"const PERF_WEEKLY_TREND = {_df_to_js(df_weekly)};",
        "",
        f"const PERF_ERROR_DIST = {_df_to_js(df_dist)};",
        "",
        f"const PERF_DATE_RANGE = {{ start: '{df['Date'].min().strftime('%Y-%m-%d')}', end: '{df['Date'].max().strftime('%Y-%m-%d')}' }};",
    ]
    
    with open(WEB_DATA_FILE, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    logger.info(f"✅ Web data saved: {WEB_DATA_FILE}")


def _dict_to_js(d):
    import json
    return json.dumps(d, ensure_ascii=False)

def _df_to_js(df):
    import json
    if df.empty:
        return '[]'
    records = df.to_dict('records')
    for r in records:
        for k, v in r.items():
            if isinstance(v, (np.integer,)): r[k] = int(v)
            elif isinstance(v, (np.floating,)): r[k] = round(float(v), 2) if not np.isnan(v) else None
            elif isinstance(v, (np.bool_,)): r[k] = bool(v)
    return json.dumps(records, ensure_ascii=False)


# ==========================================
# MAIN
# ==========================================
def main():
    print("=" * 60)
    print("📊 MODEL PERFORMANCE EVALUATION")
    print(f"   Date: {CURRENT_DATE}")
    print("=" * 60)
    
    df = load_master_data()
    if df.empty:
        print("❌ No data available. Run forecast first.")
        return
    
    # Filter to shift-based rows only for shift analysis
    has_shifts = df['Shift'].isin(['MORNING', 'EVENING']).sum()
    print(f"\n   Shift-based rows: {has_shifts:,} / {len(df):,}")
    
    print("\n📊 Building Excel report...")
    build_excel(df)
    
    print("🌐 Exporting web data...")
    export_web_data(df)
    
    # Print summary
    overall = calc_metrics(df)
    df_shift = analyze_by_shift(df)
    
    print(f"\n{'='*60}")
    print(f"📋 SUMMARY:")
    print(f"   Overall Hit Rate: {overall['Hit_Rate']}%")
    print(f"   Overall MAE: {overall['MAE']} khách")
    print(f"   Overall MAPE: {overall['MAPE']}%")
    
    if not df_shift.empty:
        print(f"\n   🌅 CA SÁNG vs 🌙 CA TỐI:")
        for _, r in df_shift.iterrows():
            print(f"      {r['Shift_Label']}: Hit={r['Hit_Rate']}%, MAE={r['MAE']}, MAPE={r['MAPE']}%")
    
    # Weekday vs Weekend
    wd_data = df[~df['Is_Weekend']]
    we_data = df[df['Is_Weekend']]
    wd_m = calc_metrics(wd_data)
    we_m = calc_metrics(we_data)
    print(f"\n   📅 NGÀY THƯỜNG vs CUỐI TUẦN:")
    print(f"      Ngày thường: Hit={wd_m['Hit_Rate']}%, MAE={wd_m['MAE']}, MAPE={wd_m['MAPE']}%")
    print(f"      Cuối tuần:   Hit={we_m['Hit_Rate']}%, MAE={we_m['MAE']}, MAPE={we_m['MAPE']}%")
    
    print(f"\n💾 Output: {OUTPUT_FILE}")
    print(f"💾 Web data: {WEB_DATA_FILE}")
    print("=" * 60)


if __name__ == '__main__':
    main()
